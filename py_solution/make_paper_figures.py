"""
论文配图批量生成脚本

输出：paper_assets/generated/paper_fig_*.png
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle, Rectangle
from matplotlib.patches import FancyArrowPatch
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import openpyxl

import config as cfg
from simulation import (
    missile_position, simulate_single_bomb,
    get_target_keypoints, _bomb_coverage_mask,
)

plt.rcParams.update({
    "font.sans-serif": ["SimHei", "Microsoft YaHei", "PingFang SC", "sans-serif"],
    "axes.unicode_minus": False,
    "figure.dpi": 150,
    "axes.grid": True,
    "grid.alpha": 0.25,
    "grid.linewidth": 0.5,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "lines.linewidth": 1.4,
})

C_MISSILE = "#C0392B"
C_DRONE   = "#2E6F9E"
C_TARGET  = "#4C4C4C"
C_FAKE    = "#8A8A8A"
C_SMOKE   = "#7F8C8D"
C_ACCENT  = "#D68910"
C_OK      = "#27AE60"


def out_path(name):
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "paper_assets", "generated", name)
    p = os.path.abspath(p)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    return p


# ============================================================
# 图1 场景初始布局
# ============================================================
def fig_scenario():
    fig = plt.figure(figsize=(9, 6.5))
    ax = fig.add_subplot(111, projection="3d")
    M = cfg.MISSILES_INIT
    D = cfg.DRONES_INIT
    ax.scatter(M[:,0], M[:,1], M[:,2], c=C_MISSILE, marker="^", s=90, label="来袭导弹", zorder=5)
    for i, n in enumerate(["M1","M2","M3"]):
        ax.text(M[i,0]+400, M[i,1], M[i,2], n, color=C_MISSILE, fontsize=9)
    ax.scatter(D[:,0], D[:,1], D[:,2], c=C_DRONE, marker="o", s=70, label="无人机", zorder=5)
    for i, n in enumerate(["FY1","FY2","FY3","FY4","FY5"]):
        ax.text(D[i,0]+400, D[i,1], D[i,2], n, color=C_DRONE, fontsize=9)
    ax.scatter(*cfg.FAKE_TARGET, c=C_FAKE, marker="x", s=80, label="假目标(原点)", zorder=5)
    ax.scatter(*cfg.TARGET_CENTER, c=C_TARGET, marker="s", s=60, label="真目标", zorder=5)
    for i in range(3):
        ax.plot([M[i,0], 0], [M[i,1], 0], [M[i,2], 0],
                color=C_MISSILE, linewidth=0.6, linestyle="--", alpha=0.4)
    ax.set_xlabel("X (m)"); ax.set_ylabel("Y (m)"); ax.set_zlabel("Z (m)")
    ax.set_title("场景初始布局（导弹、无人机、真/假目标）", fontsize=12)
    ax.legend(loc="upper left", frameon=False, fontsize=9)
    ax.view_init(elev=18, azim=-60)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_1_scenario.png"), dpi=180)
    plt.close(fig)
    print("  fig_scenario OK")


# ============================================================
# 图2 视锥线遮蔽判据几何示意（2D 模式图）
# ============================================================
def fig_view_cone():
    fig, ax = plt.subplots(figsize=(10, 6))
    M = np.array([1.0, 7.0])
    S = np.array([5.5, 4.0])
    P_T = np.array([8.0, 2.0])
    R = 1.0
    vec = P_T - M; norm = np.linalg.norm(vec)
    alpha = np.arcsin(R / np.linalg.norm(S - M))  # 半顶角
    # 云团圆
    circ = plt.Circle(S, R, color=C_DRONE, fill=True, alpha=0.30, linewidth=1.5)
    ax.add_patch(circ)
    circ_edge = plt.Circle(S, R, color=C_DRONE, fill=False, linewidth=1.5)
    ax.add_patch(circ_edge)
    # 视锥母线：从 M 画到 S 处的两条切线（用半顶角近似）
    axis = (P_T - M) / np.linalg.norm(P_T - M)
    perp = np.array([-axis[1], axis[0]])
    L = np.linalg.norm(S - M)
    # 视锥左/右母线
    p_L = S + L * np.tan(alpha) * perp + (np.linalg.norm(P_T - S)) * axis
    p_R = S - L * np.tan(alpha) * perp + (np.linalg.norm(P_T - S)) * axis
    ax.plot([M[0], p_L[0]], [M[1], p_L[1]], color=C_DRONE, linewidth=1.2, linestyle="--", alpha=0.8, label="视锥母线")
    ax.plot([M[0], p_R[0]], [M[1], p_R[1]], color=C_DRONE, linewidth=1.2, linestyle="--", alpha=0.8)
    # 视锥轴
    ax.annotate("", xy=P_T, xytext=M,
                arrowprops=dict(arrowstyle="->", color="black", lw=1.0, linestyle="--", alpha=0.5))
    # 真目标
    target_w, target_h = 0.6, 1.4
    rect = plt.Rectangle((P_T[0]-target_w/2, P_T[1]-target_h/2), target_w, target_h,
                         color=C_TARGET, alpha=0.5)
    ax.add_patch(rect)
    # 关键点
    for sign in [-1, 1]:
        for y_ in [-0.3, 0, 0.3]:
            ax.scatter([P_T[0] + sign*target_w/2], [P_T[1] + y_],
                       c=C_ACCENT, s=30, marker="o", zorder=5)
    # 视线
    for sign in [-1, 1]:
        for y_ in [-0.3, 0, 0.3]:
            ax.plot([M[0], P_T[0] + sign*target_w/2], [M[1], P_T[1] + y_],
                    color="gray", linewidth=0.5, alpha=0.4, linestyle=":")
    # 标注
    label_kw = dict(fontsize=12, fontweight='bold')
    ax.text(M[0]-0.2, M[1]+0.3, "A 导弹 M", color=C_MISSILE, **label_kw)
    ax.text(S[0]-0.3, S[1]+0.3, "云团 S", color=C_DRONE, **label_kw)
    ax.text(P_T[0]+0.1, P_T[1]+0.9, "真目标 G", color=C_TARGET, **label_kw)
    ax.text(S[0]+0.15, S[1]+0.05, "R", color=C_DRONE, fontsize=10, fontweight='bold')
    # 半顶角标注
    ax.text(3.0, 5.5, r"$\alpha$ (半顶角)", color=C_DRONE, fontsize=12, style='italic')
    # 公式标注
    ax.text(2.5, 0.8, r"$\sin\alpha = R \,/\, \|M - S\|$", fontsize=12,
            bbox=dict(boxstyle="round,pad=0.3", facecolor="#FFF8E1", edgecolor=C_ACCENT))
    # 条件框
    ax.text(0.2, 0.5, "完全遮蔽 ⟺ ∀关键点 K_i: γ_i ≤ α",
            fontsize=10, transform=ax.transAxes,
            bbox=dict(boxstyle="round,pad=0.3", facecolor="#E8F4F8", edgecolor=C_DRONE))
    ax.set_xlim(0, 9.5); ax.set_ylim(0, 8.5)
    ax.set_aspect("equal")
    ax.set_xlabel("X (等比示意)"); ax.set_ylabel("Y (等比示意)")
    ax.set_title("视锥线遮蔽判据几何示意（2D 模式图）", fontsize=13, pad=10)
    ax.legend(loc="upper right", frameon=True, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_2_view_cone.png"), dpi=180)
    plt.close(fig)
    print("  fig_view_cone OK")


# ============================================================
# 图3 多机多云团协同遮蔽机理（2D 概念示意图，关键点错列）
# ============================================================
def fig_multi_cloud():
    fig, ax = plt.subplots(figsize=(12, 6.5))

    # 目标圆柱（左端）—— 用矩形表示
    target_x = 0.5
    target_w = 1.0
    target_h = 3.0
    rect = plt.Rectangle((target_x, 0), target_w, target_h,
                         color=C_TARGET, alpha=0.4, ec="black", lw=1.2)
    ax.add_patch(rect)
    ax.text(target_x + target_w/2, target_h + 0.25, "真目标 (圆柱纵截面)",
            ha="center", fontsize=12, color=C_TARGET, fontweight='bold')

    # 关键点（6个，2列错开避免与云重叠）
    # 左侧列（紧贴目标）：K1上、K3中、K5下
    # 右侧列（靠近云团位置）：K2上、K4中、K6下
    keypoints = [
        (target_x + target_w, 2.4, "K1"),   # 紧贴目标 上
        (4.2, 2.4, "K2"),                   # 靠近云1 上
        (target_x + target_w, 1.5, "K3"),   # 紧贴目标 中
        (4.2, 1.5, "K4"),                   # 靠近云2 中
        (target_x + target_w, 0.6, "K5"),   # 紧贴目标 下
        (4.2, 0.6, "K6"),                   # 靠近云3 下
    ]
    for (x_, y_, lab) in keypoints:
        ax.scatter([x_], [y_], c=C_ACCENT, s=55, zorder=6, edgecolor='black', linewidth=0.6)
        # 标注：左列放在更左、右列放在更右，避免被云挡住
        if x_ < 1.5:
            ax.text(x_-0.25, y_, lab, fontsize=11, fontweight='bold', va='center', ha='right')
        else:
            ax.text(x_+0.30, y_, lab, fontsize=11, fontweight='bold', va='center', ha='left')

    # 视线（导弹→关键点，灰色虚线）
    missile_x = 9.5
    missile_y = 1.5
    ax.scatter([missile_x], [missile_y], c=C_MISSILE, marker="^", s=220, zorder=5)
    ax.text(missile_x+0.2, missile_y+0.15, "导弹 M", fontsize=12, color=C_MISSILE, fontweight='bold')
    for (x_, y_, lab) in keypoints:
        ax.plot([missile_x, x_], [missile_y, y_], color="gray", linewidth=0.4, alpha=0.25, linestyle=":")

    # 三朵云（垂直方向分开更多，避免重叠）
    # 云1：挡上半 K1, K2
    c1 = plt.Circle((5.0, 2.4), 0.6, color=C_SMOKE, alpha=0.6, ec=C_DRONE, lw=1.5)
    ax.add_patch(c1)
    ax.text(5.0, 3.30, "云 1\n(挡 K1, K2)", ha="center", fontsize=10, color=C_DRONE, fontweight='bold')

    # 云2：挡中段 K3, K4
    c2 = plt.Circle((5.0, 1.5), 0.6, color="#5DADE2", alpha=0.6, ec="#1B4F72", lw=1.5)
    ax.add_patch(c2)
    ax.text(5.0, 2.40, "云 2\n(挡 K3, K4)", ha="center", fontsize=10, color="#1B4F72", fontweight='bold')

    # 云3：挡下半 K5, K6
    c3 = plt.Circle((5.0, 0.6), 0.6, color="#48C9B0", alpha=0.6, ec="#117A65", lw=1.5)
    ax.add_patch(c3)
    ax.text(5.0, 0.0, "云 3\n(挡 K5, K6)", ha="center", fontsize=10, color="#117A65", fontweight='bold')

    # 视线遮挡（彩色粗线）—— 导弹到被各云挡住的视线
    blocked_map = {"K1":C_DRONE, "K2":C_DRONE, "K3":"#1B4F72", "K4":"#1B4F72", "K5":"#117A65", "K6":"#117A65"}
    for (x_, y_, lab) in keypoints:
        col = blocked_map[lab]
        ax.plot([missile_x, x_], [missile_y, y_], color=col, linewidth=2.0, alpha=0.7)

    # "被挡住"标记（绿色圈）
    for (x_, y_, _) in keypoints:
        ax.scatter([x_], [y_], s=320, facecolor='none', edgecolor='green', linewidth=2.0, zorder=4)

    # 公式框
    ax.text(0.5, 3.95, r"完全遮蔽 ⟺ $\forall$ 关键点 K_i: 至少一朵云挡住对应视线",
            fontsize=11, fontweight='bold',
            bbox=dict(boxstyle="round,pad=0.4", facecolor="#FFF8E1", edgecolor=C_ACCENT, lw=1.0))

    # 视线方向箭头
    ax.annotate("", xy=(missile_x-0.3, missile_y), xytext=(missile_x+1.0, missile_y),
                arrowprops=dict(arrowstyle="->", color="black", lw=1.2))
    ax.text(missile_x-1.3, missile_y+0.25, "视线方向", fontsize=10, color="black")

    ax.set_xlim(-1.0, 11)
    ax.set_ylim(-0.7, 4.2)
    ax.set_aspect("equal")
    ax.set_xlabel("沿视线方向 (m, 等比示意)", fontsize=11)
    ax.set_ylabel("目标高度方向 (m, 等比示意)", fontsize=11)
    ax.set_title("多机多云团协同（互补）遮蔽机理示意", fontsize=13, pad=10)
    ax.legend(handles=[
        plt.scatter([0], [0], c=C_MISSILE, marker="^", s=100, label="导弹"),
        plt.scatter([0], [0], c=C_TARGET, marker="s", s=100, label="真目标（截面）"),
        plt.scatter([0], [0], c=C_ACCENT, s=50, label="关键点 K_i"),
        plt.scatter([0], [0], s=200, facecolor='none', edgecolor='green', lw=2, label="被云挡住"),
    ], loc="lower right", frameon=True, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_3_multi_cloud.png"), dpi=180)
    plt.close(fig)
    print("  fig_multi_cloud OK")


# ============================================================
# 图4 问题1距离时序
# ============================================================
def fig_p1_timeline():
    drone_init = cfg.DRONES_INIT[0]
    theta = cfg.P1_DRONE_THETA
    speed = cfg.P1_DRONE_SPEED
    tr = cfg.P1_RELEASE_TIME
    td = cfg.P1_DETONATION_DELAY
    te = tr + td
    te_end = te + cfg.EFFECTIVE_DURATION
    ts = np.arange(0, 15, 0.001)
    M_pos = np.array([missile_position(t, 0) for t in ts])
    dist_mt = np.linalg.norm(M_pos - cfg.TARGET_CENTER, axis=1)
    direction = np.array([np.cos(theta), np.sin(theta), 0.0])
    active = (ts >= te) & (ts <= te_end)
    smoke_pos = np.full((len(ts), 3), np.nan)
    if np.any(active):
        FY_rel = drone_init + speed * direction * tr
        bx = FY_rel[0] + speed * direction[0] * td
        by = FY_rel[1] + speed * direction[1] * td
        bh = FY_rel[2] - 0.5 * cfg.G * td**2
        t_a = ts[active]
        smoke_pos[active, 0] = bx
        smoke_pos[active, 1] = by
        smoke_pos[active, 2] = bh - cfg.SMOKE_SINK_SPEED * (t_a - te)
    dist_ms = np.linalg.norm(M_pos - smoke_pos, axis=1)
    valid = ~np.isnan(dist_ms)
    i_min = np.nanargmin(dist_ms)
    t_min, d_min = ts[i_min], dist_ms[i_min]
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(ts, dist_mt, color=C_MISSILE, linewidth=1.6, label="导弹—真目标距离")
    ax.plot(ts, dist_ms, color=C_DRONE, linewidth=1.6, label="导弹—云心距离")
    ax.axhline(cfg.EFFECTIVE_RADIUS, color=C_ACCENT, linewidth=1.0, linestyle=":",
               label=f"有效遮蔽半径 {cfg.EFFECTIVE_RADIUS:.0f} m")
    ax.axvline(te, color="black", linewidth=0.8, linestyle="--", alpha=0.6)
    ax.text(te, ax.get_ylim()[1]*0.92, f" 起爆 t={te:.1f}s", fontsize=8, va="top")
    ax.set_xlabel("t (s)"); ax.set_ylabel("距离 (m)")
    ax.set_title(f"问题1距离时序：release={tr}s, delay={td}s "
                 f"(下沉{cfg.SMOKE_SINK_SPEED}m/s, 速度{cfg.DRONE_SPEED_MIN:.0f}-{cfg.DRONE_SPEED_MAX:.0f}m/s)",
                 fontsize=10)
    ax.legend(loc="center right", frameon=False, fontsize=9)
    axins = ax.inset_axes([0.10, 0.14, 0.30, 0.42])
    win = (ts > t_min - 1.5) & (ts < t_min + 1.5)
    axins.plot(ts[win], dist_ms[win], color=C_DRONE, linewidth=1.4)
    axins.axhline(cfg.EFFECTIVE_RADIUS, color=C_ACCENT, linewidth=1.0, linestyle=":")
    axins.scatter([t_min], [d_min], color=C_DRONE, s=18, zorder=5)
    axins.annotate(f"最近 {d_min:.2f} m\n(阈值{cfg.EFFECTIVE_RADIUS:.0f}m, 差{d_min-cfg.EFFECTIVE_RADIUS:+.2f}m)",
                    xy=(t_min, d_min), xytext=(0.55, 0.75), textcoords="axes fraction",
                    fontsize=7.5, ha="left",
                    arrowprops=dict(arrowstyle="-", color="#666666", linewidth=0.6))
    axins.set_xlabel("t (s)", fontsize=7)
    axins.tick_params(labelsize=7)
    axins.grid(alpha=0.25, linewidth=0.4)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_4_p1_timeline.png"), dpi=180)
    plt.close(fig)
    print("  fig_p1_timeline OK")


# ============================================================
# 图5 烟幕弹投放→平抛→起爆→下沉 全流程
# ============================================================
def fig_bomb_trajectory():
    drone_init = np.array([17800., 0., 1800.])
    theta = np.radians(178.2); v = 83.8
    tr = 0.30; td = 2.91
    direction = np.array([np.cos(theta), np.sin(theta), 0.0])
    FY_rel = drone_init + v * direction * tr
    t_drop = np.linspace(0, td, 60)
    drop_x = FY_rel[0] + v * direction[0] * t_drop
    drop_y = FY_rel[1] + v * direction[1] * t_drop
    drop_z = FY_rel[2] - 0.5 * cfg.G * t_drop**2
    burst = np.array([drop_x[-1], drop_y[-1], drop_z[-1]])
    te = tr + td
    t_sink = np.linspace(0, 20, 80)
    sink_x = np.full_like(t_sink, burst[0])
    sink_y = np.full_like(t_sink, burst[1])
    sink_z = burst[2] - cfg.SMOKE_SINK_SPEED * t_sink
    t_fly = np.linspace(0, tr+td+2, 50)
    fly_x = drone_init[0] + v * direction[0] * t_fly
    fly_y = drone_init[1] + v * direction[1] * t_fly
    fly_z = np.full_like(t_fly, drone_init[2])
    fig = plt.figure(figsize=(9, 6))
    ax = fig.add_subplot(111, projection="3d")
    ax.scatter([0],[0],[0], c=C_FAKE, marker="x", s=60, label="假目标")
    r, h = cfg.TARGET_RADIUS, cfg.TARGET_HEIGHT
    z_arr = np.linspace(0, h, 8); th_arr = np.linspace(0, 2*np.pi, 16)
    Z, TH = np.meshgrid(z_arr, th_arr)
    Xc = cfg.TARGET_CENTER[0] + r*np.cos(TH)
    Yc = cfg.TARGET_CENTER[1] + r*np.sin(TH)
    ax.plot_surface(Xc, Yc, Z, color=C_TARGET, alpha=0.4, edgecolor='none')
    ax.plot(fly_x, fly_y, fly_z, color=C_DRONE, linewidth=1.2, label="FY1 航迹")
    ax.plot(drop_x, drop_y, drop_z, color=C_ACCENT, linewidth=1.6, label="烟幕弹平抛段")
    ax.scatter([burst[0]], [burst[1]], [burst[2]], c="red", marker="*", s=120, label="起爆点")
    ax.plot(sink_x, sink_y, sink_z, color=C_SMOKE, linewidth=1.6, label=f"云团下沉 (v_s={cfg.SMOKE_SINK_SPEED}m/s)")
    for tt in [0, 5, 10, 15]:
        idx = int(tt / 20 * (len(t_sink)-1))
        cx, cy, cz = sink_x[idx], sink_y[idx], sink_z[idx]
        u_, v_ = np.meshgrid(np.linspace(0, np.pi, 8), np.linspace(0, 2*np.pi, 12))
        xs = cx + cfg.EFFECTIVE_RADIUS * np.sin(u_) * np.cos(v_)
        ys = cy + cfg.EFFECTIVE_RADIUS * np.sin(u_) * np.sin(v_)
        zs = cz + cfg.EFFECTIVE_RADIUS * np.cos(u_)
        ax.plot_surface(xs, ys, zs, color=C_SMOKE, alpha=0.20, edgecolor='none')
    ax.set_xlabel("X (m)"); ax.set_ylabel("Y (m)"); ax.set_zlabel("Z (m)")
    ax.set_title("烟幕弹投放→平抛→起爆→下沉 全流程示意（示例策略）", fontsize=12)
    ax.legend(loc="upper left", frameon=False, fontsize=9)
    ax.view_init(elev=20, azim=-60)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_5_bomb_trajectory.png"), dpi=180)
    plt.close(fig)
    print("  fig_bomb_trajectory OK")


# ============================================================
# 图6 问题2最优策略
# ============================================================
def fig_p2_strategy():
    drone_init = cfg.DRONES_INIT[0]
    theta = np.radians(178.2); v = 83.8
    tr = 0.30; td = 2.91
    te = tr + td
    direction = np.array([np.cos(theta), np.sin(theta), 0.0])
    FY_rel = drone_init + v * direction * tr
    burst = np.array([FY_rel[0] + v*direction[0]*td, FY_rel[1] + v*direction[1]*td, FY_rel[2] - 0.5*cfg.G*td**2])
    dt = 0.01
    ts = np.arange(te, te+cfg.EFFECTIVE_DURATION, dt)
    kp = get_target_keypoints(180, 5)
    mask = _bomb_coverage_mask(drone_init, theta, v, tr, td, 0, kp, dt, ts[-1]+1)
    i0 = int(te/dt); i1 = min(len(mask), i0 + int(cfg.EFFECTIVE_DURATION/dt))
    win = mask[i0:i1]
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    ax = axes[0]
    ax.scatter(0, 0, c=C_FAKE, marker="x", s=100, label="假目标")
    th = np.linspace(0, 2*np.pi, 50)
    ax.plot(cfg.TARGET_CENTER[0]+cfg.TARGET_RADIUS*np.cos(th),
            cfg.TARGET_CENTER[1]+cfg.TARGET_RADIUS*np.sin(th), color=C_TARGET, label="真目标圆周")
    ax.scatter([drone_init[0]], [drone_init[1]], c=C_DRONE, marker="o", s=80, label="FY1 起点")
    ax.scatter([FY_rel[0]], [FY_rel[1]], c=C_ACCENT, marker="s", s=80, label="投放点")
    ax.scatter([burst[0]], [burst[1]], c="red", marker="*", s=200, label="起爆点", zorder=5)
    ax.annotate("", xy=(FY_rel[0], FY_rel[1]), xytext=(drone_init[0], drone_init[1]),
                arrowprops=dict(arrowstyle="->", color=C_DRONE, lw=1.4))
    t_M = np.linspace(0, 30, 60)
    M1 = np.array([missile_position(t, 0) for t in t_M])
    ax.plot(M1[:,0], M1[:,1], color=C_MISSILE, linewidth=1.2, linestyle="--", alpha=0.7, label="M1 轨迹")
    ax.set_xlim(16500, 18500); ax.set_ylim(-300, 300)
    ax.set_xlabel("X (m)"); ax.set_ylabel("Y (m)")
    ax.set_title("问题2最优策略（俯视图，聚焦真目标区域）", fontsize=11)
    ax.legend(loc="upper right", frameon=False, fontsize=8)
    ax.grid(True, alpha=0.3)
    ax.set_aspect("equal", adjustable="datalim")
    ax2 = axes[1]
    ax2.fill_between(ts[:len(win)], 0, win.astype(float), step="mid",
                     color=C_OK, alpha=0.7, label="有效遮蔽区间")
    ax2.set_xlabel("t (s)"); ax2.set_ylabel("是否遮蔽 (0/1)")
    ax2.set_title(f"问题2最优策略 遮蔽时序 (T_eff={win.sum()*dt:.3f}s)", fontsize=11)
    ax2.set_ylim(-0.05, 1.1)
    ax2.legend(loc="upper right", frameon=False, fontsize=9)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_6_p2_strategy.png"), dpi=180)
    plt.close(fig)
    print("  fig_p2_strategy OK")


# ============================================================
# 图7 问题3 三弹时序
# ============================================================
def fig_p3_timing():
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), "result1.xlsx"))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    ts_release = [r[5] for r in rows]
    ts_deton = [r[5] + r[6] for r in rows]
    drone_init = cfg.DRONES_INIT[0]
    theta = np.radians(178.247); v = 83.81
    kp = get_target_keypoints(180, 5)
    fig, ax = plt.subplots(figsize=(10, 4.5))
    for i, (r, d) in enumerate(zip(ts_release, ts_deton)):
        m = _bomb_coverage_mask(drone_init, theta, v, r, d-r, 0, kp, 0.01, d+cfg.EFFECTIVE_DURATION+1)
        i0 = int((d)/0.01); i1 = min(len(m), i0+int(cfg.EFFECTIVE_DURATION/0.01))
        win = m[i0:i1]
        ts = np.arange(d, d+cfg.EFFECTIVE_DURATION, 0.01)[:len(win)]
        ax.fill_between(ts, i+0.1, i+0.9, where=win, step="mid",
                        color=C_OK, alpha=0.7)
        ax.text(d, i+0.5, f"弹{i+1}\n起爆{d:.2f}s",
                ha="center", va="center", fontsize=8)
    ax.set_xlabel("t (s)"); ax.set_ylabel("烟幕弹编号")
    ax.set_yticks([0.5, 1.5, 2.5])
    ax.set_yticklabels(["弹1", "弹2", "弹3"])
    ax.set_title("问题3：三枚烟幕弹各自的起爆-下沉时序（绿色=有效遮蔽区间）", fontsize=11)
    ax.set_xlim(0, 30)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_7_p3_timing.png"), dpi=180)
    plt.close(fig)
    print("  fig_p3_timing OK")


# ============================================================
# 图8 PSO收敛曲线
# ============================================================
def fig_pso_convergence():
    from pso import PSO
    from problem2 import Problem2Objective
    kp = get_target_keypoints(180, 5)
    obj = Problem2Objective(cfg.DRONES_INIT[0], kp, dt=cfg.SEARCH_DT)
    bounds = [(2.73, 3.53), (cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX),
              (0.0, 15.0), (0.0, 6.0)]
    pso = PSO(obj, bounds, n_particles=80, max_iter=60, maximize=True, verbose=False)
    _, f_opt = pso.optimize()
    hist = pso.history
    values = [-h for h in hist]
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(np.arange(len(values)), values, color=C_DRONE, linewidth=1.4)
    ax.set_xlabel("迭代次数"); ax.set_ylabel("最优目标值 (s)")
    ax.set_title(f"问题2 PSO 收敛曲线（最终 T_eff={f_opt:.3f}s）", fontsize=11)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_8_pso_conv.png"), dpi=180)
    plt.close(fig)
    print("  fig_pso_convergence OK")


# ============================================================
# 图9 关键点采样数目收敛性
# ============================================================
def fig_keypoint_convergence():
    theta = np.radians(178.2); v = 83.8
    tr = 0.30; td = 2.91
    drone_init = cfg.DRONES_INIT[0]
    n_c_list = [10, 20, 36, 50, 72, 100, 144, 180, 240, 360, 540, 720]
    vals = []
    for n_c in n_c_list:
        kp = get_target_keypoints(n_c, max(1, n_c//36))
        T = simulate_single_bomb(drone_init, theta, v, tr, td, 0, kp, 0.01, 30.0)
        vals.append(T)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(n_c_list, vals, "o-", color=C_DRONE, linewidth=1.4, markersize=6)
    ax.axvline(100, color=C_ACCENT, linewidth=0.8, linestyle="--", alpha=0.7, label="n_c=100 参考线")
    ax.set_xlabel("每圆周采样点数 n_c"); ax.set_ylabel("有效遮蔽时长 T_eff (s)")
    ax.set_title("关键点采样数目对遮蔽时长的收敛性（问题2）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_9_kp_convergence.png"), dpi=180)
    plt.close(fig)
    print("  fig_keypoint_convergence OK")
    return list(zip(n_c_list, vals))


# ============================================================
# 图10 时间步长收敛性
# ============================================================
def fig_dt_convergence():
    theta = np.radians(178.2); v = 83.8
    tr = 0.30; td = 2.91
    drone_init = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(180, 5)
    dt_list = [0.05, 0.02, 0.01, 0.005, 0.002, 0.001, 0.0005]
    vals = []
    for dt in dt_list:
        T = simulate_single_bomb(drone_init, theta, v, tr, td, 0, kp, dt, 30.0)
        vals.append(T)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(dt_list, vals, "s-", color=C_DRONE, linewidth=1.4, markersize=6)
    ax.set_xscale("log")
    ax.set_xlabel("时间步长 Δt (s, log)"); ax.set_ylabel("有效遮蔽时长 T_eff (s)")
    ax.set_title("时间步长对遮蔽时长的收敛性（问题2）", fontsize=11)
    ax.grid(True, alpha=0.3, which="both")
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_10_dt_convergence.png"), dpi=180)
    plt.close(fig)
    print("  fig_dt_convergence OK")
    return list(zip(dt_list, vals))


# ============================================================
# 图11 云团下沉速度影响
# ============================================================
def fig_sink_speed():
    theta = np.radians(178.2); v = 83.8
    tr = 0.30; td = 2.91
    drone_init = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(180, 5)
    vs_list = [1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 5.0]
    vals = []
    for vs in vs_list:
        old = cfg.SMOKE_SINK_SPEED
        cfg.SMOKE_SINK_SPEED = vs
        T = simulate_single_bomb(drone_init, theta, v, tr, td, 0, kp, 0.01, 30.0)
        vals.append(T)
        cfg.SMOKE_SINK_SPEED = old
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(vs_list, vals, "o-", color=C_DRONE, linewidth=1.4, markersize=6)
    ax.axvline(2.5, color=C_ACCENT, linewidth=0.8, linestyle="--", alpha=0.7, label="本题 v_s=2.5")
    ax.axvline(3.0, color="#888888", linewidth=0.8, linestyle="--", alpha=0.7, label="对照 v_s=3.0")
    ax.set_xlabel("云团下沉速度 v_s (m/s)"); ax.set_ylabel("有效遮蔽时长 T_eff (s)")
    ax.set_title("云团下沉速度对遮蔽时长的影响（问题2）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_11_sink_speed.png"), dpi=180)
    plt.close(fig)
    print("  fig_sink_speed OK")
    return list(zip(vs_list, vals))


# ============================================================
# 图12 无人机速度上限影响
# ============================================================
def fig_drone_speed():
    """无人机速度上限敏感性：每个 v_max 下若原始最优 v=83.8 ≤ v_max 则保持 v=83.8，
    否则取 v=v_max 并扫描；展示"上限收窄"对最优解的影响。"""
    theta = np.radians(178.2)
    tr = 0.30; td = 2.91
    drone_init = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(180, 5)
    v_opt = 83.8  # 问题2原始最优速度
    vmax_list = [85, 90, 100, 110, 120, 130, 140, 150]
    vals = []
    for vmax in vmax_list:
        v_use = min(v_opt, vmax)
        T = simulate_single_bomb(drone_init, theta, v_use, tr, td, 0, kp, 0.01, 30.0)
        vals.append(T)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(vmax_list, vals, "o-", color=C_DRONE, linewidth=1.4, markersize=6)
    ax.axvline(120, color=C_ACCENT, linewidth=0.8, linestyle="--", alpha=0.7, label="本题 v_max=120")
    ax.set_xlabel("无人机速度上限 v_max (m/s)"); ax.set_ylabel("有效遮蔽时长 T_eff (s)")
    ax.set_title("无人机速度上限对遮蔽时长的影响（问题2, v=min(83.8, v_max)）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_12_drone_speed.png"), dpi=180)
    plt.close(fig)
    print("  fig_drone_speed OK")
    return list(zip(vmax_list, vals))


# ============================================================
# 图A1 投放时刻敏感性（问题2最优解附近扰动）
# ============================================================
def fig_release_time():
    theta = np.radians(178.2); v = 83.8; td = 2.91
    drone_init = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(180, 5)
    tr_list = [0.05, 0.10, 0.20, 0.30, 0.40, 0.50, 0.60, 0.80, 1.00, 1.50]
    vals = []
    for tr in tr_list:
        T = simulate_single_bomb(drone_init, theta, v, tr, td, 0, kp, 0.01, 30.0)
        vals.append(T)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(tr_list, vals, "o-", color=C_DRONE, linewidth=1.4, markersize=6)
    ax.axvline(0.30, color=C_ACCENT, linewidth=0.8, linestyle="--", alpha=0.7, label="最优 tr=0.30 s")
    ax.set_xlabel("投放时刻 $t_r$ (s)")
    ax.set_ylabel("有效遮蔽时长 $T_{eff}$ (s)")
    ax.set_title("投放时刻 $t_r$ 对有效遮蔽时长的影响（问题2）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_a1_release_time.png"), dpi=180)
    plt.close(fig)
    print("  fig_release_time OK")
    return list(zip(tr_list, vals))


# ============================================================
# 图A2 起爆延时敏感性（问题2最优解附近扰动）
# ============================================================
def fig_detonation_delay():
    theta = np.radians(178.2); v = 83.8; tr = 0.30
    drone_init = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(180, 5)
    td_list = [1.5, 2.0, 2.5, 2.8, 2.91, 3.0, 3.2, 3.5, 4.0, 5.0]
    vals = []
    for td in td_list:
        T = simulate_single_bomb(drone_init, theta, v, tr, td, 0, kp, 0.01, 30.0)
        vals.append(T)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(td_list, vals, "s-", color=C_DRONE, linewidth=1.4, markersize=6)
    ax.axvline(2.91, color=C_ACCENT, linewidth=0.8, linestyle="--", alpha=0.7, label="最优 td=2.91 s")
    ax.set_xlabel("起爆延时 $t_d$ (s)")
    ax.set_ylabel("有效遮蔽时长 $T_{eff}$ (s)")
    ax.set_title("起爆延时 $t_d$ 对有效遮蔽时长的影响（问题2）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_a2_detonation_delay.png"), dpi=180)
    plt.close(fig)
    print("  fig_detonation_delay OK")
    return list(zip(td_list, vals))


# ============================================================
# 图A3 算法稳定性：问题2 上重复运行 PSO 多次
# ============================================================
def fig_algorithm_stability():
    from pso import PSO
    from problem2 import Problem2Objective
    kp = get_target_keypoints(180, 5)
    obj = Problem2Objective(cfg.DRONES_INIT[0], kp, dt=cfg.SEARCH_DT)
    bounds = [(2.73, 3.53), (cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX),
              (0.0, 15.0), (0.0, 6.0)]
    n_runs = 8
    finals = []
    for run_id in range(n_runs):
        np.random.seed(run_id * 17 + 1)  # 主进程内确定性
        pso = PSO(obj, bounds, n_particles=80, max_iter=60, maximize=True,
                  verbose=False, n_workers=1)  # 单进程让 seed 可控
        _, f_opt = pso.optimize()
        finals.append(f_opt)
    fig, ax = plt.subplots(figsize=(6.5, 4))
    ax.bar(range(1, n_runs+1), finals, color=C_DRONE, alpha=0.75, edgecolor='black', linewidth=0.4)
    ax.axhline(np.mean(finals), color=C_ACCENT, linewidth=1.0, linestyle="--",
               label=f"均值 {np.mean(finals):.3f} s")
    ax.axhline(np.max(finals), color=C_OK, linewidth=1.0, linestyle=":",
               label=f"最大值 {np.max(finals):.3f} s")
    ax.set_xlabel("运行序号 (不同随机种子)")
    ax.set_ylabel("最优 $T_{eff}$ (s)")
    ax.set_title(f"PSO 在问题 2 上的稳定性（{n_runs} 次独立运行）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3, axis='y')
    ax.set_ylim(min(finals) - 0.05, max(finals) + 0.05)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_a3_pso_stability.png"), dpi=180)
    plt.close(fig)
    print("  fig_algorithm_stability OK")
    return finals


# ============================================================
# 图A4 三种算法收敛曲线对比（PSO/GA/SA 在问题2 上同条件实测）
# ============================================================
def fig_algo_comparison():
    """同条件同评估预算下 PSO / GA / SA 收敛曲线"""
    from pso import PSO
    from problem2 import Problem2Objective
    from ga_sa_baselines import run_ga, run_sa
    kp = get_target_keypoints(180, 5)
    obj = Problem2Objective(cfg.DRONES_INIT[0], kp, dt=cfg.SEARCH_DT)
    bounds = [(2.73, 3.53), (cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX),
              (0.0, 15.0), (0.0, 6.0)]
    n_eval_budget = 6000  # 公平对比：每个算法最多 6000 次评估
    # PSO: 80 粒子 × 75 代 ≈ 6000
    np.random.seed(42)
    pso = PSO(obj, bounds, n_particles=80, max_iter=75, maximize=True,
              verbose=False, n_workers=1)
    _ = pso.optimize()
    pso_curve = [-h for h in pso.history]
    # GA / SA 用统一的评估预算
    ga_curve = run_ga(obj, bounds, n_eval_budget=n_eval_budget, seed=42)
    sa_curve = run_sa(obj, bounds, n_eval_budget=n_eval_budget, seed=42)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(np.arange(len(pso_curve))*80, pso_curve, color=C_DRONE, linewidth=1.4, label="PSO")
    ax.plot(np.arange(len(ga_curve)), ga_curve, color=C_ACCENT, linewidth=1.4, label="GA (实数编码)")
    ax.plot(np.arange(len(sa_curve)), sa_curve, color="#117A65", linewidth=1.4, label="SA (指数降温)")
    ax.set_xlabel("函数评估次数")
    ax.set_ylabel("当前最优 $T_{eff}$ (s)")
    ax.set_title("PSO / GA / SA 在问题 2 上的收敛曲线（相同评估预算 6000 次）", fontsize=11)
    ax.legend(frameon=False, fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_a4_algo_compare.png"), dpi=180)
    plt.close(fig)
    print("  fig_algo_comparison OK")
    return {"pso": pso_curve[-1], "ga": ga_curve[-1], "sa": sa_curve[-1]}


# ============================================================
# 图13 问题4 三机协同策略（去掉云团球避免3D重叠，只画航迹+起爆点）
# ============================================================
def fig_p4_strategy():
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), "result2.xlsx"))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    fig = plt.figure(figsize=(10, 6.5))
    ax = fig.add_subplot(111, projection="3d")
    M = cfg.MISSILES_INIT
    D = cfg.DRONES_INIT
    ax.scatter(M[0,0], M[0,1], M[0,2], c=C_MISSILE, marker="^", s=120, label="M1 (来袭)")
    ax.scatter(D[:3,0], D[:3,1], D[:3,2], c=C_DRONE, marker="o", s=80, label="FY1-3 (起点)")
    ax.scatter(*cfg.FAKE_TARGET, c=C_FAKE, marker="x", s=80, label="假目标(原点)")
    # 真目标（浅色）
    r, h = cfg.TARGET_RADIUS, cfg.TARGET_HEIGHT
    z_arr = np.linspace(0, h, 6); th_arr = np.linspace(0, 2*np.pi, 12)
    Z, TH = np.meshgrid(z_arr, th_arr)
    Xc = cfg.TARGET_CENTER[0] + r*np.cos(TH)
    Yc = cfg.TARGET_CENTER[1] + r*np.sin(TH)
    ax.plot_surface(Xc, Yc, Z, color=C_TARGET, alpha=0.3, edgecolor='none')
    colors = [C_DRONE, "#48C9B0", "#D68910"]
    for i, (r_, col) in enumerate(zip(rows, colors)):
        theta = r_[1]; v = r_[3]; tr = r_[4]; td = r_[5]
        direction = np.array([np.cos(theta), np.sin(theta), 0.0])
        rel = D[i] + v * direction * tr
        burst = np.array([rel[0]+v*direction[0]*td, rel[1]+v*direction[1]*td, rel[2]-0.5*cfg.G*td**2])
        t_fly = np.linspace(0, tr+td+0.5, 30)
        fx = D[i,0] + v*direction[0]*t_fly
        fy = D[i,1] + v*direction[1]*t_fly
        fz = np.full_like(t_fly, D[i,2])
        ax.plot(fx, fy, fz, color=col, linewidth=1.4, alpha=0.85, label=f"FY{i+1} 航迹")
        # 起爆点（带标签）
        ax.scatter([burst[0]], [burst[1]], [burst[2]], c=col, marker="*", s=180, zorder=5)
        ax.text(burst[0], burst[1], burst[2]+30, f"FY{i+1}起爆", color=col, fontsize=9, fontweight='bold')
    # M1 轨迹
    t_M = np.linspace(0, 25, 50)
    M1 = np.array([missile_position(t, 0) for t in t_M])
    ax.plot(M1[:,0], M1[:,1], M1[:,2], color=C_MISSILE, linewidth=1.4, linestyle="--", alpha=0.7, label="M1 飞行轨迹")
    ax.set_xlabel("X (m)", fontsize=10); ax.set_ylabel("Y (m)", fontsize=10); ax.set_zlabel("Z (m)", fontsize=10)
    ax.set_title("问题4：三机协同最优策略（航迹+起爆点, 不画云团球以避免遮挡）", fontsize=12)
    ax.legend(loc="upper left", frameon=True, fontsize=8)
    ax.view_init(elev=22, azim=-55)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_13_p4_strategy.png"), dpi=180)
    plt.close(fig)
    print("  fig_p4_strategy OK")


# ============================================================
# 图14 问题5 五机多弹策略（去掉云团球，俯视+侧视双视角）
# ============================================================
def fig_p5_strategy():
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), "result3.xlsx"))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    # 双图：左=俯视图（X-Y），右=侧视图（X-Z）
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))
    M = cfg.MISSILES_INIT
    D = cfg.DRONES_INIT
    # 颜色：按目标导弹
    missile_color = {0: C_MISSILE, 1: "#1B4F72", 2: "#117A65"}
    missile_label = {0: "M1", 1: "M2", 2: "M3"}
    # === 左图：俯视（X-Y）===
    ax = axes[0]
    ax.scatter(0, 0, c=C_FAKE, marker="x", s=120, label="假目标")
    th = np.linspace(0, 2*np.pi, 50)
    ax.plot(cfg.TARGET_CENTER[0]+cfg.TARGET_RADIUS*np.cos(th),
            cfg.TARGET_CENTER[1]+cfg.TARGET_RADIUS*np.sin(th),
            color=C_TARGET, linewidth=1.5, label="真目标")
    # 3枚导弹
    for k in range(3):
        t_M = np.linspace(0, 30, 40)
        Mk = np.array([missile_position(t, k) for t in t_M])
        ax.plot(Mk[:,0], Mk[:,1], color=C_MISSILE, linewidth=1.0, linestyle="--", alpha=0.4)
        ax.scatter([Mk[0,0]], [Mk[0,1]], c=C_MISSILE, marker="^", s=80)
        ax.text(Mk[0,0]+200, Mk[0,1], f"M{k+1}", color=C_MISSILE, fontsize=10, fontweight='bold')
    # 5架无人机起爆点（按目标着色）
    for r_ in rows:
        di = int(r_[0][2]) - 1
        theta = r_[1]; v = r_[3]; tr = r_[6]; td = r_[7]
        mi = int(r_[5][1]) - 1
        direction = np.array([np.cos(theta), np.sin(theta), 0.0])
        rel = D[di] + v * direction * tr
        burst = np.array([rel[0]+v*direction[0]*td, rel[1]+v*direction[1]*td])
        col = missile_color.get(mi, C_DRONE)
        ax.scatter([burst[0]], [burst[1]], c=col, marker="*", s=100, zorder=5)
    ax.set_xlabel("X (m)"); ax.set_ylabel("Y (m)")
    ax.set_title("问题5 起爆点分布 (俯视, X-Y)", fontsize=11)
    ax.grid(True, alpha=0.3)
    ax.set_aspect("equal", adjustable="datalim")
    ax.legend(loc="upper right", frameon=True, fontsize=8)

    # === 右图：侧视（X-Z）===
    ax = axes[1]
    ax.scatter(0, 0, c=C_FAKE, marker="x", s=120, label="假目标")
    # 真目标（侧面投影：矩形）
    rect = plt.Rectangle((cfg.TARGET_CENTER[0]-cfg.TARGET_RADIUS, 0),
                          2*cfg.TARGET_RADIUS, cfg.TARGET_HEIGHT,
                          color=C_TARGET, alpha=0.5, label="真目标")
    ax.add_patch(rect)
    # 3枚导弹轨迹
    for k in range(3):
        t_M = np.linspace(0, 30, 40)
        Mk = np.array([missile_position(t, k) for t in t_M])
        ax.plot(Mk[:,0], Mk[:,2], color=C_MISSILE, linewidth=1.0, linestyle="--", alpha=0.4)
    # 5架无人机起爆点（按目标着色）
    for r_ in rows:
        di = int(r_[0][2]) - 1
        theta = r_[1]; v = r_[3]; tr = r_[6]; td = r_[7]
        mi = int(r_[5][1]) - 1
        direction = np.array([np.cos(theta), np.sin(theta), 0.0])
        rel = D[di] + v * direction * tr
        burst = np.array([rel[0]+v*direction[0]*td,
                          rel[2]-0.5*cfg.G*td**2 - cfg.SMOKE_SINK_SPEED*0])  # 起爆点z
        col = missile_color.get(mi, C_DRONE)
        ax.scatter([burst[0]], [burst[1]], c=col, marker="*", s=100, zorder=5)
    # 图例
    handles = [plt.scatter([0], [0], c=C_MISSILE, marker="*", s=100, label=f"指向 {missile_label[0]} 的起爆点"),
               plt.scatter([0], [0], c="#1B4F72", marker="*", s=100, label=f"指向 {missile_label[1]} 的起爆点"),
               plt.scatter([0], [0], c="#117A65", marker="*", s=100, label=f"指向 {missile_label[2]} 的起爆点")]
    ax.legend(handles=handles, loc="upper right", frameon=True, fontsize=8)
    ax.set_xlabel("X (m)"); ax.set_ylabel("Z (m)")
    ax.set_title("问题5 起爆点分布 (侧视, X-Z)", fontsize=11)
    ax.grid(True, alpha=0.3)
    ax.set_xlim(0, 21000)
    ax.set_ylim(0, 2500)

    fig.suptitle("问题5：五机多弹多导弹协同最优策略起爆点分布", fontsize=13, y=1.02)
    fig.tight_layout()
    fig.savefig(out_path("paper_fig_14_p5_strategy.png"), dpi=180)
    plt.close(fig)
    print("  fig_p5_strategy OK")


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    print("=== 生成论文配图 ===")
    fig_scenario()
    fig_view_cone()
    fig_multi_cloud()
    fig_p1_timeline()
    fig_bomb_trajectory()
    fig_p2_strategy()
    fig_p3_timing()
    fig_pso_convergence()
    fig_keypoint_convergence()
    fig_dt_convergence()
    fig_sink_speed()
    fig_drone_speed()
    fig_p4_strategy()
    fig_p5_strategy()
    # 敏感性增强章节新增图
    data_kp = fig_release_time()
    data_td = fig_detonation_delay()
    finals = fig_algorithm_stability()
    compare = fig_algo_comparison()
    print("---")
    print(f"PSO 稳定性: mean={np.mean(finals):.4f} std={np.std(finals):.4f} max={np.max(finals):.4f}")
    print(f"算法对比(同等预算6000次评估): PSO={compare['pso']:.3f} GA={compare['ga']:.3f} SA={compare['sa']:.3f}")
    print("=== 全部完成 ===")
