"""
运行所有5个问题 - 使用改进的优化方法
"""
import sys, os, time
import numpy as np

os.chdir(os.path.dirname(os.path.abspath(__file__)))

import config as cfg
from simulation import (
    simulate_single_bomb, simulate_multi_bomb_single_drone,
    simulate_multi_drone_multi_bomb, get_target_keypoints
)
from pso import PSO


def find_feasible_detonation(drone_init, missile_idx=0, n_samples=50000, dt=0.005):
    """
    通过大规模随机采样寻找可行的起爆位置

    关键洞察: 烟幕云团需要放在导弹→目标的视线上
    目标在y=200，导弹在y=0，所以烟幕应有y≈100-300的偏移
    """
    kp_small = get_target_keypoints(36, 0)  # 先用少量关键点快速筛选
    kp_full = get_target_keypoints(360, 0)

    M1_vx = cfg.MISSILE_SPEED * abs(cfg.MISSILES_DIR[missile_idx][0])
    M1_start_x = cfg.MISSILES_INIT[missile_idx][0]
    dx, dy, dz = drone_init
    v_min, v_max = cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX

    best_val = 0
    best_params = None

    print(f"  随机采样 {n_samples} 个候选...")
    found_any = 0

    for i in range(n_samples):
        if (i + 1) % 10000 == 0:
            print(f"    进度: {i+1}/{n_samples}, 找到: {found_any}, 最优: {best_val:.4f}s")

        # 采样: 在可行空间内随机采样
        # 烟幕位置应在导弹和目标的视线附近
        # x: M1初始位置到原点之间
        # y: 0到400之间(覆盖目标y=200)
        # z: 500到1800之间
        x = np.random.uniform(10000, 17800)
        y = np.random.uniform(-50, 400)
        z = np.random.uniform(300, 1750)

        # 可行性检查: 无人机能否在导弹到达前到达此位置?
        drone_to_pt_h = np.sqrt((x - dx)**2 + (y - dy)**2)
        if drone_to_pt_h < 1:
            continue

        # 烟幕弹需要下落的高度
        if z >= dz:
            continue

        fall_dist = dz - z
        td = np.sqrt(2 * fall_dist / cfg.G)

        # 无人机飞行时间
        # 尝试不同速度
        for v_test in [v_max, 115, 110, 100, v_min]:
            tr = drone_to_pt_h / v_test - td
            if tr < 0:
                continue

            total_time = tr + td
            M1_arrival = (M1_start_x - x) / M1_vx
            if M1_arrival <= 0 or total_time > M1_arrival:
                continue

            # 可行! 计算航向
            theta = np.arctan2(y - dy, x - dx)

            # 快速评估遮蔽效果
            val = simulate_single_bomb(
                drone_init, theta, v_test, tr, td,
                missile_idx, kp_small, dt, 30.0
            )

            if val > 0.01:
                found_any += 1
                # 用完整关键点集精确评估
                val_full = simulate_single_bomb(
                    drone_init, theta, v_test, tr, td,
                    missile_idx, kp_full, dt, 30.0
                )

                if val_full > best_val:
                    best_val = val_full
                    best_params = {
                        'pos': (x, y, z),
                        'theta': theta,
                        'speed': v_test,
                        'release_time': tr,
                        'delay': td,
                    }
                    print(f"    *** 新最优: {best_val:.4f}s "
                          f"pos=({x:.0f},{y:.0f},{z:.0f}) "
                          f"θ={np.degrees(theta):.1f}° v={v_test:.1f}m/s "
                          f"tr={tr:.3f}s td={td:.3f}s")
            break  # 只尝试第一组可行的速度

    return best_params, best_val


def run_problem1():
    """问题1: 固定参数"""
    print("\n" + "="*60)
    print("问题1")
    print("="*60)
    di = cfg.DRONES_INIT[0]
    t = simulate_single_bomb(di, np.pi, 120., 1.2, 3.2, 0, dt=cfg.DT_FINE, t_total=30.)
    print(f"参数: θ=π, v=120m/s, 投放=1.2s, 延时=3.2s")
    print(f"有效遮蔽时长: {t:.4f} s")
    return t


def run_problem2():
    """问题2: 单机单弹优化"""
    print("\n" + "="*60)
    print("问题2: FY1单机单弹最优策略")
    print("="*60)

    di = cfg.DRONES_INIT[0]

    # 阶段1: 随机采样找到可行区域
    print("阶段1: 随机采样搜索...")
    t0 = time.time()
    params, val = find_feasible_detonation(di, 0, n_samples=30000)
    print(f"  采样耗时: {time.time()-t0:.1f}s")

    if params is None:
        print("随机采样未找到有效解，尝试扩大的PSO...")
        # Fallback to PSO with wide bounds
        bounds = [(0.0, 1.0), (80, 120), (0, 15), (0.5, 10)]
        kp = get_target_keypoints(360, 0)

        def obj(x):
            return simulate_single_bomb(di, x[0], x[1], x[2], x[3], 0, kp, 0.005, 30.)

        pso = PSO(obj, bounds, n_particles=500, max_iter=200, maximize=True, verbose=True)
        x_opt, f_opt = pso.optimize()
        theta, speed, rt, dd = x_opt
    else:
        print(f"  采样最优: {val:.4f}s")
        # 阶段2: 在最优解附近用PSO精化
        print("阶段2: PSO局部精化...")
        theta_c, speed_c, rt_c, dd_c = params['theta'], params['speed'], params['release_time'], params['delay']

        delta = 0.05
        bounds = [
            (max(0.01, theta_c - delta), min(1.5, theta_c + delta)),
            (max(80, speed_c - 15), min(120, speed_c + 15)),
            (max(0, rt_c - 1), rt_c + 1),
            (max(0.1, dd_c - 1), dd_c + 1),
        ]
        kp = get_target_keypoints(360, 0)

        def obj(x):
            return simulate_single_bomb(di, x[0], x[1], x[2], x[3], 0, kp, 0.005, 30.)

        pso = PSO(obj, bounds, n_particles=200, max_iter=100, maximize=True, verbose=True)
        x_opt, f_opt = pso.optimize()
        theta, speed, rt, dd = x_opt

        if f_opt < val:  # PSO didn't improve
            theta, speed, rt, dd = theta_c, speed_c, rt_c, dd_c
            f_opt = val

    d = np.array([np.cos(theta), np.sin(theta), 0.])
    rp = di + speed * d * rt
    dp = rp + speed * d * dd; dp[2] -= 0.5 * 9.8 * dd**2

    print(f"\n最终结果:")
    print(f"  θ={theta:.4f}rad({np.degrees(theta):.1f}°), v={speed:.1f}m/s")
    print(f"  投放t={rt:.4f}s, 延时={dd:.4f}s")
    print(f"  投放点:({rp[0]:.0f},{rp[1]:.0f},{rp[2]:.0f})")
    print(f"  起爆点:({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
    print(f"  最优遮蔽时长: {f_opt:.4f} s")

    return np.array([theta, speed, rt, dd]), f_opt


def run_problem3():
    """问题3: 单机三弹"""
    print("\n" + "="*60)
    print("问题3: FY1三弹最优策略")
    print("="*60)

    di = cfg.DRONES_INIT[0]

    # 先用问题2的方法找单弹最优，然后扩展到三弹
    print("阶段1: 先找单弹最优...")
    params, val_single = find_feasible_detonation(di, 0, n_samples=20000)

    if params:
        theta0 = params['theta']
        speed0 = params['speed']
        rt0 = params['release_time']
        dd0 = params['delay']
        print(f"  单弹最优: {val_single:.4f}s, θ={np.degrees(theta0):.1f}°, v={speed0:.1f}")
    else:
        theta0, speed0, rt0, dd0 = 0.3, 120, 5, 3

    # 阶段2: PSO优化三弹
    print("阶段2: PSO优化三弹策略...")
    kp = get_target_keypoints(360, 10)
    bounds = [
        (theta0 - 0.2, theta0 + 0.2),
        (max(80, speed0-10), min(120, speed0+10)),
        (max(0, rt0 - 3), rt0 + 2),
        (1, 5), (1, 5),
        (max(0.1, dd0-2), dd0+3),
        (max(0.1, dd0-2), dd0+3),
        (max(0.1, dd0-2), dd0+3),
    ]

    def obj(x):
        rt = np.array([x[2], x[2]+x[3], x[2]+x[3]+x[4]])
        dd = np.array([x[5], x[6], x[7]])
        return simulate_multi_bomb_single_drone(di, x[0], x[1], rt, dd,
                                                 np.array([0,0,0]), kp, 0.005, 30.)

    pso = PSO(obj, bounds, n_particles=300, max_iter=150, maximize=True, verbose=True)
    x, f = pso.optimize()

    rt = np.array([x[2], x[2]+x[3], x[2]+x[3]+x[4]])
    dd = np.array([x[5], x[6], x[7]])
    d = np.array([np.cos(x[0]), np.sin(x[0]), 0.])

    print(f"\nθ={np.degrees(x[0]):.1f}°, v={x[1]:.1f}m/s")
    for i in range(3):
        rp = di + x[1]*d*rt[i]; dp = rp + x[1]*d*dd[i]; dp[2]-=0.5*9.8*dd[i]**2
        print(f"  弹{i+1}: 投放t={rt[i]:.4f}s 延时={dd[i]:.4f}s 起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
    print(f"总遮蔽: {f:.4f}s")

    # Save
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题3"
    h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","投放时间s","延时s",
         "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
    for c,hv in enumerate(h,1): ws.cell(1,c,hv)
    for i in range(3):
        rp = di + x[1]*d*rt[i]; dp2 = rp + x[1]*d*dd[i]; dp2[2]-=0.5*9.8*dd[i]**2
        rw=i+2
        ws.cell(rw,1,"FY1"); ws.cell(rw,2,round(x[0],6)); ws.cell(rw,3,round(np.degrees(x[0]),4))
        ws.cell(rw,4,round(x[1],2)); ws.cell(rw,5,i+1)
        ws.cell(rw,6,round(rt[i],4)); ws.cell(rw,7,round(dd[i],4))
        ws.cell(rw,8,round(rp[0],2)); ws.cell(rw,9,round(rp[1],2)); ws.cell(rw,10,round(rp[2],2))
        ws.cell(rw,11,round(dp2[0],2)); ws.cell(rw,12,round(dp2[1],2)); ws.cell(rw,13,round(dp2[2],2))
        if i==0: ws.cell(rw,14,round(f,4))
    wb.save("result1.xlsx"); print("已保存 result1.xlsx")
    return x, f


def run_problem4():
    """问题4: 三机各一弹"""
    print("\n" + "="*60)
    print("问题4: FY1/FY2/FY3各一弹协同")
    print("="*60)

    results = []
    names = ['FY1', 'FY2', 'FY3']

    for idx in range(3):
        print(f"\n--- 优化 {names[idx]} ---")
        di = cfg.DRONES_INIT[idx]
        params, val = find_feasible_detonation(di, 0, n_samples=20000)
        if params:
            results.append({**params, 'name': names[idx], 'drone_idx': idx})
            print(f"  {names[idx]}: {val:.4f}s, θ={np.degrees(params['theta']):.1f}°, "
                  f"v={params['speed']:.1f}m/s")
        else:
            print(f"  {names[idx]}: 未找到有效解!")

    if not results:
        print("没有任何无人机找到有效解!"); return None, 0

    # 计算协同遮蔽
    print("\n阶段2: 协同仿真...")
    drone_params = []
    for r in results:
        drone_params.append({
            'drone_init': cfg.DRONES_INIT[r['drone_idx']],
            'theta': r['theta'], 'speed': r['speed'],
            'release_times': np.array([r['release_time']]),
            'detonation_delays': np.array([r['delay']]),
            'missile_indices': [0],
        })

    total, per = simulate_multi_drone_multi_bomb(drone_params, 0.005, 35.)
    print(f"\n协同遮蔽总时长: {total:.4f}s")

    # Save
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题4"
    h = ["无人机","航向角rad","航向角°","速度m/s","投放时间s","延时s",
         "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
    for c,hv in enumerate(h,1): ws.cell(1,c,hv)
    for i, r in enumerate(results):
        d = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.])
        di = cfg.DRONES_INIT[r['drone_idx']]
        rp = di + r['speed']*d*r['release_time']
        dp = rp + r['speed']*d*r['delay']; dp[2]-=0.5*9.8*r['delay']**2
        rw=i+2
        ws.cell(rw,1,r['name']); ws.cell(rw,2,round(r['theta'],6)); ws.cell(rw,3,round(np.degrees(r['theta']),4))
        ws.cell(rw,4,round(r['speed'],2)); ws.cell(rw,5,round(r['release_time'],4)); ws.cell(rw,6,round(r['delay'],4))
        ws.cell(rw,7,round(rp[0],2)); ws.cell(rw,8,round(rp[1],2)); ws.cell(rw,9,round(rp[2],2))
        ws.cell(rw,10,round(dp[0],2)); ws.cell(rw,11,round(dp[1],2)); ws.cell(rw,12,round(dp[2],2))
        if i==0: ws.cell(rw,13,round(total,4))
    wb.save("result2.xlsx"); print("已保存 result2.xlsx")
    return results, total


def run_problem5():
    """问题5: 五机多弹多导弹"""
    print("\n" + "="*60)
    print("问题5: 5机协同多弹多导弹")
    print("="*60)

    names = ['FY1','FY2','FY3','FY4','FY5']
    order = cfg.INTERCEPT_ORDER
    n_drones = 5
    n_bombs = 3

    all_results = []
    for di in range(n_drones):
        print(f"\n--- 优化 {names[di]} (拦截: {order[names[di]]}) ---")
        drone_init = cfg.DRONES_INIT[di]
        od = order[names[di]]

        # 为每枚弹单独找最优起爆位置
        bomb_params = []
        for bomb_idx, missile_idx in enumerate(od):
            print(f"  弹{bomb_idx+1} -> M{missile_idx+1}:")
            params, val = find_feasible_detonation(drone_init, missile_idx, n_samples=15000)
            if params:
                bomb_params.append({**params, 'val': val, 'missile_idx': missile_idx})
                print(f"    遮蔽{val:.4f}s, pos=({params['pos'][0]:.0f},{params['pos'][1]:.0f},{params['pos'][2]:.0f})")

        if bomb_params:
            # 统一航向和速度(使用第一枚弹的)
            theta = bomb_params[0]['theta']
            speed = bomb_params[0]['speed']
            rt = np.array([p['release_time'] for p in bomb_params])
            # 调整投放时间为递增(最小间隔1s)
            for j in range(1, len(rt)):
                if rt[j] < rt[j-1] + 1:
                    rt[j] = rt[j-1] + 1
            dd = np.array([p['delay'] for p in bomb_params])
            mi = np.array([p['missile_idx'] for p in bomb_params])

            all_results.append({
                'name': names[di], 'drone_idx': di,
                'theta': theta, 'speed': speed,
                'release_times': rt, 'delays': dd,
                'missile_indices': mi,
            })
            print(f"  {names[di]}: θ={np.degrees(theta):.1f}°, v={speed:.1f}m/s, "
                  f"投放={rt}, 延时={dd}")

    if not all_results:
        print("没有找到有效解!"); return None, 0

    # 协同仿真
    print("\n协同仿真...")
    drone_params_list = []
    for r in all_results:
        drone_params_list.append({
            'drone_init': cfg.DRONES_INIT[r['drone_idx']],
            'theta': r['theta'], 'speed': r['speed'],
            'release_times': r['release_times'],
            'detonation_delays': r['delays'],
            'missile_indices': list(r['missile_indices']),
        })

    total, per = simulate_multi_drone_multi_bomb(drone_params_list, 0.005, 40.)
    print(f"\n总遮蔽时长: {total:.4f}s (M1:{per[0]:.2f}, M2:{per[1]:.2f}, M3:{per[2]:.2f})")

    # Save
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题5"
    h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","目标导弹","投放时间s","延时s",
         "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
    for c,hv in enumerate(h,1): ws.cell(1,c,hv)
    rw=2
    for r in all_results:
        d = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.])
        di = cfg.DRONES_INIT[r['drone_idx']]
        for j in range(len(r['release_times'])):
            rp = di + r['speed']*d*r['release_times'][j]
            dp = rp + r['speed']*d*r['delays'][j]; dp[2]-=0.5*9.8*r['delays'][j]**2
            ws.cell(rw,1,r['name']); ws.cell(rw,2,round(r['theta'],6))
            ws.cell(rw,3,round(np.degrees(r['theta']),4)); ws.cell(rw,4,round(r['speed'],2))
            ws.cell(rw,5,j+1); ws.cell(rw,6,f"M{r['missile_indices'][j]+1}")
            ws.cell(rw,7,round(r['release_times'][j],4)); ws.cell(rw,8,round(r['delays'][j],4))
            ws.cell(rw,9,round(rp[0],2)); ws.cell(rw,10,round(rp[1],2)); ws.cell(rw,11,round(rp[2],2))
            ws.cell(rw,12,round(dp[0],2)); ws.cell(rw,13,round(dp[1],2)); ws.cell(rw,14,round(dp[2],2))
            if rw==2: ws.cell(rw,15,round(total,4))
            rw+=1
    wb.save("result3.xlsx"); print("已保存 result3.xlsx")
    return all_results, total


if __name__ == "__main__":
    t0 = time.time()

    r1 = run_problem1()
    t1 = time.time(); print(f"\n[P1耗时: {t1-t0:.1f}s]")

    r2 = run_problem2()
    t2 = time.time(); print(f"\n[P2耗时: {t2-t1:.1f}s]")

    r3 = run_problem3()
    t3 = time.time(); print(f"\n[P3耗时: {t3-t2:.1f}s]")

    r4 = run_problem4()
    t4 = time.time(); print(f"\n[P4耗时: {t4-t3:.1f}s]")

    r5 = run_problem5()
    t5 = time.time(); print(f"\n[P5耗时: {t5-t4:.1f}s]")

    print("\n" + "="*70)
    print("                    最终结果汇总 (C题)")
    print("="*70)
    print(f"  问题1: {r1:.4f} s")
    print(f"  问题2: {r2[1]:.4f} s")
    print(f"  问题3: {r3[1]:.4f} s  -> result1.xlsx")
    print(f"  问题4: {r4[1]:.4f} s  -> result2.xlsx")
    print(f"  问题5: {r5[1]:.4f} s  -> result3.xlsx")
    print(f"\n  总耗时: {t5-t0:.1f}s ({(t5-t0)/60:.1f}min)")
    print("="*70)
    print("\n参考值 (2025 A题原参数):")
    print("  问题1: 1.3915s / 问题2: 4.5960s / 问题3: 7.6500s")
    print("  问题4: 11.7540s / 问题5: 38.0600s")
    print("  C题参数: 下沉2.5m/s(原3.0), 速度80-120m/s(原70-140)")
    print("  C题P1: 投放1.2s/延时3.2s(原1.5/3.6)")
