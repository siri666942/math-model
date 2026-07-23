"""
问题4: 利用FY1、FY2、FY3各投放1枚烟幕干扰弹，实施对M1的干扰
输出结果到 result2.xlsx

优化变量(12维):
[theta1, theta2, theta3, speed1, speed2, speed3,
 release1, release2, release3, delay1, delay2, delay3]

求解策略参考了2025年国奖论文(cumcm25003等)的"3-1-1拆解+边界收缩"思路：
一次性对12维做冷启动全空间PSO代价太大(实测单次评估约0.9s，500粒子x300代要接近
40小时)，容易陷入局部最优。改成两阶段：
  阶段1: 把"3架无人机各投1弹"拆解成3个独立的"1架无人机投1弹"问题分别求解，
        每架无人机只优化自己对M1的遮蔽时长，互不知道彼此在干什么，快速拿到
        一个"够用"的基线解。
  阶段2: 以阶段1每架无人机各自的最优参数为中心，收缩12维的搜索边界，在这个
        小得多的范围内做真正的12维联合PSO精修——目标函数是三机协同后的并集
        遮蔽时长(会考虑互补/错峰效应)，不是阶段1那种各自为战的目标。
这样搜索空间大幅收窄，评估次数也能相应减少，同时不放弃"用真实协同目标函数做
最后把关"这一步。
"""
import numpy as np
import openpyxl
from config import (
    DRONES_INIT, DRONE_SPEED_MIN, DRONE_SPEED_MAX, DT, T_TOTAL,
)
from simulation import simulate_multi_drone_multi_bomb, get_target_keypoints
from pso import PSO, local_polish
from final_solve import search_best_detonation

# 阶段1兜底用: 如果三维空间采样也没找到可行解，才退回猜角度窗口的PSO再试一次
PSO_SWARM_P4_STAGE1_FALLBACK = 150
PSO_ITER_P4_STAGE1_FALLBACK = 60
STAGE1_N_FAST = 20000  # search_best_detonation 阶段1的采样点数

# 阶段2: 12维联合精修的PSO参数(搜索范围已经收窄，不需要跟冷启动时一样大的预算)
PSO_SWARM_P4_STAGE2 = 200
PSO_ITER_P4_STAGE2 = 100

# 各无人机朝向真目标(0,200,0)的方位角窗口——只用于阶段1兜底PSO的搜索范围，
# 正常路径下阶段1走search_best_detonation(三维空间采样反解方向)，不依赖这个猜测窗口。
# 之前的教训：FY2/FY3若直接靠这个"指向目标"猜出来的窗口搜索，会把真正的最优方向
# 排除在外(实测跟国奖论文报告的FY2/FY3最优方向对不上)，导致阶段1两架都搜出0。
THETA_WINDOWS = [
    (2.73, 3.53),      # FY1, 目标方位约179.4°
    (-3.44, -2.64),    # FY2, 目标方位约-174.3°
    (2.25, 3.05),      # FY3, 目标方位约151.9°
]

N_DRONES = 3


class Problem4Objective:
    """阶段2用: 12维联合目标函数(三机协同、取并集)，模块级可pickle"""

    def __init__(self, n_drones):
        self.n_drones = n_drones

    def __call__(self, x):
        theta = x[0:3]
        speed = x[3:6]
        release_times = x[6:9]
        delays = x[9:12]

        drone_params = []
        for i in range(self.n_drones):
            drone_params.append({
                'drone_init': DRONES_INIT[i],
                'theta': theta[i],
                'speed': speed[i],
                'release_times': np.array([release_times[i]]),
                'detonation_delays': np.array([delays[i]]),
                'missile_indices': [0],  # 全部针对M1
            })

        total_time, per_missile = simulate_multi_drone_multi_bomb(
            drone_params, dt=DT, t_total=T_TOTAL
        )
        return total_time


class Stage1DroneObjective:
    """阶段1用: 单架无人机独自对M1的遮蔽时长，模块级可pickle"""

    def __init__(self, drone_idx):
        self.drone_idx = drone_idx

    def __call__(self, x):
        theta, speed, release_time, delay = x
        drone_params = [{
            'drone_init': DRONES_INIT[self.drone_idx],
            'theta': theta,
            'speed': speed,
            'release_times': np.array([release_time]),
            'detonation_delays': np.array([delay]),
            'missile_indices': [0],
        }]
        total_time, _ = simulate_multi_drone_multi_bomb(
            drone_params, dt=DT, t_total=T_TOTAL
        )
        return total_time


def _narrow_bounds(center, full_lo, full_hi, margin):
    """以center为中心收缩出一个不超过[full_lo,full_hi]的小区间"""
    lo = max(full_lo, center - margin)
    hi = min(full_hi, center + margin)
    if hi <= lo:  # 极端情况下退化保护，保证区间非空
        hi = lo + 1e-6
    return lo, hi


def solve_problem4():
    """求解问题4: 三机各一弹对M1（阶段1独立拆解 + 阶段2边界收缩联合精修）"""
    print("=" * 60)
    print("问题4: FY1/FY2/FY3各投放1枚烟幕弹对M1 (两阶段PSO优化)")
    print("=" * 60)

    drone_names = ['FY1', 'FY2', 'FY3']

    # ============================================================
    # 阶段1: 逐架无人机独立优化 (拆解成3个"1机1弹"问题)
    # 用search_best_detonation在无人机周围的三维空间里采样候选起爆点、反解方向，
    # 不预先猜角度窗口——避免猜错窗口把真正的最优方向排除在外(FY2/FY3就吃过这个亏)。
    # ============================================================
    print("\n阶段1: 逐架无人机独立优化(三维空间采样，不预设角度窗口)...")

    stage1_results = []
    for i in range(N_DRONES):
        params, f_i = search_best_detonation(DRONES_INIT[i], missile_idx=0, n_fast=STAGE1_N_FAST)

        if params is None:
            # 兜底: 三维空间采样也没找到可行解，退回猜角度窗口的PSO再试一次
            print(f"  {drone_names[i]}: 三维采样未找到可行解，退回窗口PSO兜底...")
            bounds_i = [THETA_WINDOWS[i], (DRONE_SPEED_MIN, DRONE_SPEED_MAX), (0.0, 20.0), (0.0, 20.0)]
            obj_i = Stage1DroneObjective(i)
            pso_i = PSO(obj_i, bounds_i, n_particles=PSO_SWARM_P4_STAGE1_FALLBACK,
                        max_iter=PSO_ITER_P4_STAGE1_FALLBACK, maximize=True, verbose=False)
            x_i, f_i = pso_i.optimize()
            theta_i, speed_i, rt_i, delay_i = x_i[0], x_i[1], x_i[2], x_i[3]
        else:
            theta_i = params['theta']
            speed_i = params['speed']
            rt_i = params['release_time']
            delay_i = params['delay']

        stage1_results.append({'theta': theta_i, 'speed': speed_i,
                                'release_time': rt_i, 'delay': delay_i, 'time': f_i})
        print(f"  {drone_names[i]}: θ={np.degrees(theta_i):.1f}° v={speed_i:.1f}m/s "
              f"独自遮蔽={f_i:.4f}s")

    total_stage1 = sum(r['time'] for r in stage1_results)
    print(f"\n阶段1 单机各自最优简单求和(仅作参考基线，不是最终答案): {total_stage1:.4f} s")

    # ============================================================
    # 阶段2: 以阶段1结果为中心，收缩边界后做12维联合精修
    # ============================================================
    print("\n阶段2: 12维联合精修(边界已收缩到阶段1解附近)...")

    bounds = []
    theta_margin = 0.45   # rad，比之前略放宽，给阶段2多一点纠偏空间
    speed_margin = (DRONE_SPEED_MAX - DRONE_SPEED_MIN) * 0.3
    time_margin = 3.0     # s，release/delay 的搜索余量

    # 注意: 这里clamp用的是完整角度范围(-pi,pi]，不是THETA_WINDOWS——
    # 阶段1的theta现在来自三维空间反解，可能落在THETA_WINDOWS之外(FY2/FY3就是这样)，
    # 再用旧窗口去clamp会重新把刚找到的正确方向卡掉，等于白修。
    for i in range(N_DRONES):
        bounds.append(_narrow_bounds(stage1_results[i]['theta'], -np.pi, np.pi, theta_margin))
    for i in range(N_DRONES):
        bounds.append(_narrow_bounds(stage1_results[i]['speed'], DRONE_SPEED_MIN, DRONE_SPEED_MAX, speed_margin))
    for i in range(N_DRONES):
        bounds.append(_narrow_bounds(stage1_results[i]['release_time'], 0.0, 20.0, time_margin))
    for i in range(N_DRONES):
        bounds.append(_narrow_bounds(stage1_results[i]['delay'], 0.0, 20.0, time_margin))

    objective = Problem4Objective(N_DRONES)

    print(f"\n变量维度: 12 (3×theta, 3×speed, 3×release, 3×delay)")
    print(f"粒子群规模: {PSO_SWARM_P4_STAGE2}, 迭代次数: {PSO_ITER_P4_STAGE2}")

    pso = PSO(objective, bounds, n_particles=PSO_SWARM_P4_STAGE2,
              max_iter=PSO_ITER_P4_STAGE2, maximize=True, verbose=True)
    x_opt, f_opt = pso.optimize()

    print("\n局部精修(Powell)...")
    x_polished, f_polished = local_polish(objective, x_opt, bounds)
    if f_polished > f_opt:
        print(f"  精修有提升: {f_opt:.4f}s -> {f_polished:.4f}s")
        x_opt, f_opt = x_polished, f_polished
    else:
        print(f"  精修没有提升({f_polished:.4f}s <= {f_opt:.4f}s)，保留PSO原结果")

    # 解析结果
    theta = x_opt[0:3]
    speed = x_opt[3:6]
    release_times = x_opt[6:9]
    delays = x_opt[9:12]

    print(f"\n优化结果:")
    for i in range(N_DRONES):
        direction = np.array([np.cos(theta[i]), np.sin(theta[i]), 0.0])
        release_pos = DRONES_INIT[i] + speed[i] * direction * release_times[i]
        detonation_pos = release_pos + speed[i] * direction * delays[i]
        detonation_pos[2] -= 0.5 * 9.8 * delays[i] ** 2

        print(f"\n  {drone_names[i]}:")
        print(f"    航向角θ: {theta[i]:.4f} rad ({np.degrees(theta[i]):.2f}°)")
        print(f"    飞行速度: {speed[i]:.2f} m/s")
        print(f"    投放时间: {release_times[i]:.4f} s")
        print(f"    起爆延时: {delays[i]:.4f} s")
        print(f"    投放点: ({release_pos[0]:.2f}, {release_pos[1]:.2f}, {release_pos[2]:.2f})")
        print(f"    起爆点: ({detonation_pos[0]:.2f}, {detonation_pos[1]:.2f}, {detonation_pos[2]:.2f})")

    print(f"\n  阶段1基线(各自为战简单求和): {total_stage1:.4f} s")
    print(f"  阶段2联合精修总有效遮蔽时长: {f_opt:.4f} s")

    # 保存到 result2.xlsx
    save_result2(theta, speed, release_times, delays, f_opt)

    return x_opt, f_opt


def save_result2(theta, speed, release_times, delays, total_time):
    """保存问题4结果到 result2.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问题4结果"

    headers = ["无人机", "航向角(rad)", "航向角(°)", "飞行速度(m/s)",
               "投放时间(s)", "起爆延时(s)",
               "投放点X", "投放点Y", "投放点Z",
               "起爆点X", "起爆点Y", "起爆点Z",
               "总有效遮蔽时长(s)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    drone_names = ['FY1', 'FY2', 'FY3']
    for i in range(3):
        row = i + 2
        direction = np.array([np.cos(theta[i]), np.sin(theta[i]), 0.0])
        release_pos = DRONES_INIT[i] + speed[i] * direction * release_times[i]
        detonation_pos = release_pos + speed[i] * direction * delays[i]
        detonation_pos[2] -= 0.5 * 9.8 * delays[i] ** 2

        ws.cell(row=row, column=1, value=drone_names[i])
        ws.cell(row=row, column=2, value=round(theta[i], 6))
        ws.cell(row=row, column=3, value=round(np.degrees(theta[i]), 4))
        ws.cell(row=row, column=4, value=round(speed[i], 2))
        ws.cell(row=row, column=5, value=round(release_times[i], 4))
        ws.cell(row=row, column=6, value=round(delays[i], 4))
        ws.cell(row=row, column=7, value=round(release_pos[0], 2))
        ws.cell(row=row, column=8, value=round(release_pos[1], 2))
        ws.cell(row=row, column=9, value=round(release_pos[2], 2))
        ws.cell(row=row, column=10, value=round(detonation_pos[0], 2))
        ws.cell(row=row, column=11, value=round(detonation_pos[1], 2))
        ws.cell(row=row, column=12, value=round(detonation_pos[2], 2))
        if i == 0:
            ws.cell(row=row, column=13, value=round(total_time, 4))

    filepath = "result2.xlsx"
    wb.save(filepath)
    print(f"\n结果已保存至: {filepath}")


if __name__ == "__main__":
    solve_problem4()
