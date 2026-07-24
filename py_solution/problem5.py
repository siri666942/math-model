"""
问题5: 5架无人机，每架至多投放3枚烟幕干扰弹，实施对M1、M2、M3的干扰
输出结果到 result3.xlsx

优化变量(40维):
- theta(5): 各无人机航向角
- speed(5): 各无人机飞行速度
- release_times(5×3=15): 各弹投放时间(编码为间隔)
- detonation_delays(5×3=15): 各弹起爆延时

采用分步优化: 先单机优化，再联合微调
"""
import numpy as np
import openpyxl
from config import (
    DRONES_INIT, DRONE_SPEED_MIN, DRONE_SPEED_MAX,
    BOMB_INTERVAL_MIN, INTERCEPT_ORDER, DT, T_TOTAL,
    SEARCH_N_CIRCLE, SEARCH_N_LAYERS, SEARCH_DT,
    FINAL_N_CIRCLE, FINAL_N_LAYERS, FINAL_DT,
)
from simulation import simulate_multi_drone_multi_bomb, get_target_keypoints
from pso import PSO, local_polish
from final_solve import search_best_detonation

# 阶段1三维采样的采样点数(与P4一致的量级)
STAGE1_N_FAST_P5 = 8000

# 预算已按搜索档实测吞吐重定：40维联合评估约0.2s/10进程。
# 阶段2 = 80×50约14分钟；阶段1每机 60×25。整个问题5约24分钟。
PSO_SWARM_P5 = 80
PSO_ITER_P5 = 50

N_DRONES = 5
N_BOMBS_PER_DRONE = 3
N_MISSILES = 3

# 搜索档/定稿档关键点(模块级预生成，供objective引用)
_SEARCH_KP = get_target_keypoints(SEARCH_N_CIRCLE, SEARCH_N_LAYERS)
_FINAL_KP = get_target_keypoints(FINAL_N_CIRCLE, FINAL_N_LAYERS)


class SingleDroneObjective:
    """阶段1: 单机独立优化用的目标函数对象，模块级可pickle，供PSO多进程worker调用"""

    def __init__(self, drone_idx, missile_order):
        self.drone_idx = drone_idx
        self.missile_order = missile_order

    def __call__(self, x):
        theta, speed = x[0], x[1]
        release_times = np.array([x[2], x[2] + x[3], x[2] + x[3] + x[4]])
        delays = np.array([x[5], x[6], x[7]])

        drone_params = [{
            'drone_init': DRONES_INIT[self.drone_idx],
            'theta': theta,
            'speed': speed,
            'release_times': release_times,
            'detonation_delays': delays,
            'missile_indices': self.missile_order,
        }]
        total_time, _ = simulate_multi_drone_multi_bomb(
            drone_params, dt=SEARCH_DT, t_total=T_TOTAL, target_keypoints=_SEARCH_KP
        )
        return total_time


def _joint_objective(x):
    """阶段2: 40维联合微调用的目标函数，不捕获任何局部变量，模块级可pickle"""
    idx = 0
    drone_params = []
    for drone_idx in range(N_DRONES):
        theta = x[idx]; idx += 1
        speed = x[idx]; idx += 1

        release1 = x[idx]; idx += 1
        int2 = x[idx]; idx += 1
        int3 = x[idx]; idx += 1
        release_times = np.array([release1, release1 + int2, release1 + int2 + int3])

        delays = np.array([x[idx + j] for j in range(3)]); idx += 3
        order = INTERCEPT_ORDER[['FY1', 'FY2', 'FY3', 'FY4', 'FY5'][drone_idx]]

        drone_params.append({
            'drone_init': DRONES_INIT[drone_idx],
            'theta': theta,
            'speed': speed,
            'release_times': release_times,
            'detonation_delays': delays,
            'missile_indices': order,
        })

    total_time, _ = simulate_multi_drone_multi_bomb(
        drone_params, dt=SEARCH_DT, t_total=T_TOTAL, target_keypoints=_SEARCH_KP
    )
    return total_time


def _joint_objective_final(x):
    """定稿档复算：与 _joint_objective 同解码，只是换成360关键点/dt0.005精细档"""
    idx = 0
    drone_params = []
    for drone_idx in range(N_DRONES):
        theta = x[idx]; idx += 1
        speed = x[idx]; idx += 1
        release1 = x[idx]; idx += 1
        int2 = x[idx]; idx += 1
        int3 = x[idx]; idx += 1
        release_times = np.array([release1, release1 + int2, release1 + int2 + int3])
        delays = np.array([x[idx + j] for j in range(3)]); idx += 3
        order = INTERCEPT_ORDER[['FY1', 'FY2', 'FY3', 'FY4', 'FY5'][drone_idx]]
        drone_params.append({
            'drone_init': DRONES_INIT[drone_idx], 'theta': theta, 'speed': speed,
            'release_times': release_times, 'detonation_delays': delays,
            'missile_indices': order,
        })
    total_time, _ = simulate_multi_drone_multi_bomb(
        drone_params, dt=FINAL_DT, t_total=T_TOTAL, target_keypoints=_FINAL_KP
    )
    return total_time


def solve_problem5():
    """求解问题5: 五机多弹多导弹协同策略"""
    print("=" * 60)
    print("问题5: 5架无人机协同投放烟幕弹 (PSO分步优化)")
    print("=" * 60)

    # 预定义每架无人机的最佳搜索范围
    # 基于初始位置分析
    # theta_range 以各无人机指向真目标(0,200,0)的方位角为中心留出搜索余量
    # (原范围都取在0°附近的小角度，方向朝向战场正前方而非目标，PSO永远搜不到有效遮蔽)
    drone_configs = [
        {  # FY1: 目标方位约179.4°(3.13rad)
            'theta_range': (2.73, 3.53),
            'speed_range': (DRONE_SPEED_MIN, DRONE_SPEED_MAX),
            'release_range': (0.0, 5.0),
            'delay_range': (0.0, 8.0),
        },
        {  # FY2: 目标方位约-174.3°(-3.04rad)
            'theta_range': (-3.44, -2.64),
            'speed_range': (DRONE_SPEED_MIN, DRONE_SPEED_MAX),
            'release_range': (0.0, 8.0),
            'delay_range': (0.0, 10.0),
        },
        {  # FY3: 目标方位约151.9°(2.65rad)
            'theta_range': (2.25, 3.05),
            'speed_range': (DRONE_SPEED_MIN, DRONE_SPEED_MAX),
            'release_range': (0.0, 12.0),
            'delay_range': (0.0, 12.0),
        },
        {  # FY4: 目标方位约-170.7°(-2.98rad)
            'theta_range': (-3.38, -2.58),
            'speed_range': (DRONE_SPEED_MIN, DRONE_SPEED_MAX),
            'release_range': (0.0, 15.0),
            'delay_range': (0.0, 15.0),
        },
        {  # FY5: 目标方位约170.4°(2.97rad)
            'theta_range': (2.57, 3.37),
            'speed_range': (DRONE_SPEED_MIN, DRONE_SPEED_MAX),
            'release_range': (0.0, 15.0),
            'delay_range': (0.0, 15.0),
        },
    ]

    # ============================================================
    # 阶段1: 逐架无人机单独优化
    # 改用三维空间采样(search_best_detonation，与P4同款)：对该机 INTERCEPT_ORDER 里
    # 每一枚弹各自的目标导弹采样反解最优起爆点，得到每弹的(theta,speed,release,delay)；
    # 一架机只有一个航向/速度，取各弹的中位数作为该机航向/速度，再评估实际单机遮蔽。
    # (原"角度窗口PSO"方向被人为限死，FY2~FY5恒为0；三维采样不预设窗口，能找到解。)
    # ============================================================
    print("\n阶段1: 逐架无人机单独优化(三维空间采样)...")

    drone_names = ['FY1', 'FY2', 'FY3', 'FY4', 'FY5']
    single_results = []
    for drone_idx in range(N_DRONES):
        cfg = drone_configs[drone_idx]
        order = INTERCEPT_ORDER[drone_names[drone_idx]]
        print(f"\n--- 优化 {drone_names[drone_idx]} (拦截顺序 {order}) ---")

        # 对每枚弹的目标导弹分别三维采样
        per_bomb = []
        for mi in order:
            params, _ = search_best_detonation(DRONES_INIT[drone_idx], mi,
                                               n_fast=STAGE1_N_FAST_P5)
            per_bomb.append(params)  # 可能为 None

        found = [p for p in per_bomb if p is not None]
        if found:
            theta_m = float(np.median([p['theta'] for p in found]))
            speed_m = float(np.median([p['speed'] for p in found]))
            # 各弹的投放/延时；没搜到的弹用中位数兜底，保证3枚都有值
            rel_med = float(np.median([p['release_time'] for p in found]))
            dly_med = float(np.median([p['delay'] for p in found]))
            rel = np.array([p['release_time'] if p is not None else rel_med for p in per_bomb])
            dly = np.array([p['delay'] if p is not None else dly_med for p in per_bomb])
        else:
            # 三维采样一枚都没找到：退回窗口中心当兜底种子(贡献可能为0，交给阶段2再说)
            theta_m = float(np.mean(cfg['theta_range']))
            speed_m = float(np.mean(cfg['speed_range']))
            rel = np.array([0.5, 2.0, 3.5])
            dly = np.array([3.0, 3.0, 3.0])

        # 夹进物理/绝对范围(注意theta用全范围(-π,π)，不夹cfg的猜测角度窗口——
        # 三维采样反解出的最优方向可能落在窗口外，夹回窗口会把正确方向卡掉，
        # 这正是P4踩过的坑)。release/delay用较宽的绝对范围，保证阶段2边界合法。
        theta_m = float(np.clip(theta_m, -np.pi, np.pi))
        speed_m = float(np.clip(speed_m, DRONE_SPEED_MIN, DRONE_SPEED_MAX))
        rel = np.clip(rel, 0.0, 20.0)
        dly = np.clip(dly, 0.0, 15.0)
        # 投放时间按弹序排序并强制≥1s间隔(相邻投放约束)
        rel = np.sort(rel)
        for j in range(1, len(rel)):
            if rel[j] < rel[j - 1] + BOMB_INTERVAL_MIN:
                rel[j] = rel[j - 1] + BOMB_INTERVAL_MIN

        # 用装配好的单机配置评估实际单机遮蔽时长(搜索档)，作为阶段1参考
        f_single = SingleDroneObjective(drone_idx, order)(
            np.array([theta_m, speed_m, rel[0], rel[1] - rel[0], rel[2] - rel[1],
                      dly[0], dly[1], dly[2]]))

        single_results.append({
            'theta': theta_m, 'speed': speed_m,
            'release_times': rel, 'delays': dly,
            'time': f_single,
        })
        print(f"  {drone_names[drone_idx]}: θ={np.degrees(theta_m):.1f}° v={speed_m:.1f}m/s "
              f"单机遮蔽={f_single:.4f}s")

    total_single = sum(r['time'] for r in single_results)
    print(f"\n阶段1 单机优化总时长(简单求和): {total_single:.4f} s")

    # ============================================================
    # 阶段2: 联合局部微调 (缩小搜索范围)
    # ============================================================
    print("\n阶段2: 联合局部微调...")

    # 构建40维变量，搜索范围以阶段1结果为中心。
    # 关键：变量顺序必须与 _joint_objective 的解码顺序一致——每架无人机
    # 连续8维 [theta, speed, release1, int2, int3, delay1, delay2, delay3]，
    # 一架接一架交错排列（旧代码曾按"变量类型分组"排列，与解码顺序错位，导致
    # PSO搜的是完全打乱的空间、恒为0，这里修正为交错排列）。
    # 同时用阶段1各机的解拼成一个热启动种子，避免40维纯随机初始化搜不到有效区。
    bounds_joint = []
    seed_joint = []
    for drone_idx in range(N_DRONES):
        r = single_results[drone_idx]

        # theta: 以阶段1解为中心留±0.35rad(~20°)搜索余量，clamp到全范围(-π,π)
        # (不用cfg的猜测窗口做clamp——阶段1的theta来自三维采样，可能在窗口外)
        t_center = r['theta']
        theta_lo = max(-np.pi, t_center - 0.35)
        theta_hi = min(np.pi, t_center + 0.35)
        bounds_joint.append((theta_lo, theta_hi))
        seed_joint.append(np.clip(t_center, theta_lo, theta_hi))

        # speed: ±10 m/s，clamp到[80,120]
        s_center = r['speed']
        speed_lo = max(DRONE_SPEED_MIN, s_center - 10.0)
        speed_hi = min(DRONE_SPEED_MAX, s_center + 10.0)
        bounds_joint.append((speed_lo, speed_hi))
        seed_joint.append(np.clip(s_center, speed_lo, speed_hi))

        # release1: ±1.0s，clamp到[0,20]
        rel_lo = max(0.0, r['release_times'][0] - 1.0)
        rel_hi = min(20.0, r['release_times'][0] + 1.0)
        bounds_joint.append((rel_lo, rel_hi))
        seed_joint.append(np.clip(r['release_times'][0], rel_lo, rel_hi))

        # interval 1->2
        int_center = r['release_times'][1] - r['release_times'][0]
        i2_lo = max(BOMB_INTERVAL_MIN, int_center - 0.5)
        i2_hi = max(i2_lo, int_center + 0.5)
        bounds_joint.append((i2_lo, i2_hi))
        seed_joint.append(np.clip(int_center, i2_lo, i2_hi))

        # interval 2->3
        int_center2 = r['release_times'][2] - r['release_times'][1]
        i3_lo = max(BOMB_INTERVAL_MIN, int_center2 - 0.5)
        i3_hi = max(i3_lo, int_center2 + 0.5)
        bounds_joint.append((i3_lo, i3_hi))
        seed_joint.append(np.clip(int_center2, i3_lo, i3_hi))

        # delays 1/2/3: ±1.5s，clamp到[0,15]
        for j in range(3):
            d_center = r['delays'][j]
            d_lo = max(0.0, d_center - 1.5)
            d_hi = min(15.0, d_center + 1.5)
            bounds_joint.append((d_lo, d_hi))
            seed_joint.append(np.clip(d_center, d_lo, d_hi))

    print(f"联合优化变量维度: {len(bounds_joint)}")
    pso_joint = PSO(_joint_objective, bounds_joint, n_particles=PSO_SWARM_P5,
                    max_iter=PSO_ITER_P5, maximize=True, verbose=True,
                    seed_positions=[np.array(seed_joint)])
    x_opt_joint, f_opt_joint = pso_joint.optimize()

    print("\n局部精修(Powell)...")
    x_polished, f_polished = local_polish(_joint_objective, x_opt_joint, bounds_joint)
    if f_polished > f_opt_joint:
        print(f"  精修有提升: {f_opt_joint:.4f}s -> {f_polished:.4f}s")
        x_opt_joint, f_opt_joint = x_polished, f_polished
    else:
        print(f"  精修没有提升({f_polished:.4f}s <= {f_opt_joint:.4f}s)，保留PSO原结果")

    # 定稿档复算：搜索用180关键点/dt0.01，最优解用360关键点/dt0.005复算上报
    f_final = _joint_objective_final(x_opt_joint)
    print(f"  定稿档复算(360关键点/dt{FINAL_DT}): {f_opt_joint:.4f}s(搜索档) -> {f_final:.4f}s(定稿)")
    f_opt_joint = f_final

    print(f"\n联合优化总有效遮蔽时长: {f_opt_joint:.4f} s")

    # ============================================================
    # 解析并输出结果
    # ============================================================
    idx = 0
    final_results = []
    for drone_idx in range(N_DRONES):
        theta = x_opt_joint[idx]; idx += 1
        speed = x_opt_joint[idx]; idx += 1
        release1 = x_opt_joint[idx]; idx += 1
        int2 = x_opt_joint[idx]; idx += 1
        int3 = x_opt_joint[idx]; idx += 1
        release_times = np.array([release1, release1+int2, release1+int2+int3])
        delays = np.array([x_opt_joint[idx+j] for j in range(3)]); idx += 3
        order = INTERCEPT_ORDER[['FY1','FY2','FY3','FY4','FY5'][drone_idx]]
        final_results.append({
            'theta': theta, 'speed': speed,
            'release_times': release_times, 'delays': delays,
            'order': order,
        })

    drone_names = ['FY1', 'FY2', 'FY3', 'FY4', 'FY5']
    print(f"\n{'='*60}")
    print("最终优化结果:")
    print(f"{'='*60}")

    for i, r in enumerate(final_results):
        direction = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.0])
        print(f"\n{drone_names[i]}: θ={r['theta']:.4f}rad({np.degrees(r['theta']):.1f}°), "
              f"v={r['speed']:.2f}m/s")
        print(f"  拦截顺序: {r['order']}")
        for j in range(3):
            release_pos = DRONES_INIT[i] + r['speed'] * direction * r['release_times'][j]
            detonation_pos = release_pos + r['speed'] * direction * r['delays'][j]
            detonation_pos[2] -= 0.5 * 9.8 * r['delays'][j] ** 2
            print(f"  弹{j+1}: 投放t={r['release_times'][j]:.4f}s, 延时={r['delays'][j]:.4f}s, "
                  f"起爆点=({detonation_pos[0]:.1f},{detonation_pos[1]:.1f},{detonation_pos[2]:.1f})")

    print(f"\n总有效遮蔽时长: {f_opt_joint:.4f} s")

    # 保存
    save_result3(final_results, f_opt_joint)

    return final_results, f_opt_joint


def save_result3(final_results, total_time):
    """保存问题5结果到 result3.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问题5结果"

    headers = ["无人机", "航向角(rad)", "航向角(°)", "飞行速度(m/s)",
               "烟幕弹编号", "目标导弹", "投放时间(s)", "起爆延时(s)",
               "投放点X", "投放点Y", "投放点Z",
               "起爆点X", "起爆点Y", "起爆点Z",
               "总有效遮蔽时长(s)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    drone_names = ['FY1', 'FY2', 'FY3', 'FY4', 'FY5']
    row = 2
    for i, r in enumerate(final_results):
        direction = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.0])
        for j in range(3):
            missile_label = f"M{r['order'][j]+1}"
            release_pos = DRONES_INIT[i] + r['speed'] * direction * r['release_times'][j]
            detonation_pos = release_pos + r['speed'] * direction * r['delays'][j]
            detonation_pos[2] -= 0.5 * 9.8 * r['delays'][j] ** 2

            ws.cell(row=row, column=1, value=drone_names[i])
            ws.cell(row=row, column=2, value=round(r['theta'], 6))
            ws.cell(row=row, column=3, value=round(np.degrees(r['theta']), 4))
            ws.cell(row=row, column=4, value=round(r['speed'], 2))
            ws.cell(row=row, column=5, value=j + 1)
            ws.cell(row=row, column=6, value=missile_label)
            ws.cell(row=row, column=7, value=round(r['release_times'][j], 4))
            ws.cell(row=row, column=8, value=round(r['delays'][j], 4))
            ws.cell(row=row, column=9, value=round(release_pos[0], 2))
            ws.cell(row=row, column=10, value=round(release_pos[1], 2))
            ws.cell(row=row, column=11, value=round(release_pos[2], 2))
            ws.cell(row=row, column=12, value=round(detonation_pos[0], 2))
            ws.cell(row=row, column=13, value=round(detonation_pos[1], 2))
            ws.cell(row=row, column=14, value=round(detonation_pos[2], 2))
            if i == 0 and j == 0:
                ws.cell(row=row, column=15, value=round(total_time, 4))
            row += 1

    filepath = "result3.xlsx"
    wb.save(filepath)
    print(f"\n结果已保存至: {filepath}")


if __name__ == "__main__":
    solve_problem5()
