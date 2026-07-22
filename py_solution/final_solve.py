"""
最终求解 - 使用优化的快速评估
"""
import numpy as np
import time, sys
import config as cfg
from simulation import (
    simulate_single_bomb, simulate_multi_bomb_single_drone,
    simulate_multi_drone_multi_bomb, get_target_keypoints
)

# 预生成不同精度的关键点
KP_FAST = get_target_keypoints(36, 0)      # 72个点快速评估
KP_MEDIUM = get_target_keypoints(180, 0)    # 360个点
KP_FULL = get_target_keypoints(360, 10)     # 完整


def fast_eval_1bomb(drone_init, theta, speed, rt, dd, mi=0):
    """快速单弹评估"""
    return simulate_single_bomb(drone_init, theta, speed, rt, dd, mi, KP_FAST, 0.02, 25.)


def full_eval_1bomb(drone_init, theta, speed, rt, dd, mi=0):
    """精确单弹评估"""
    return simulate_single_bomb(drone_init, theta, speed, rt, dd, mi, KP_FULL, 0.005, 30.)


def fast_eval_3bombs(drone_init, theta, speed, rt, dd, mi):
    """快速三弹评估"""
    return simulate_multi_bomb_single_drone(drone_init, theta, speed, rt, dd, mi, KP_FAST, 0.02, 25.)


def full_eval_3bombs(drone_init, theta, speed, rt, dd, mi):
    """精确三弹评估"""
    return simulate_multi_bomb_single_drone(drone_init, theta, speed, rt, dd, mi, KP_FULL, 0.005, 30.)


def search_best_detonation(drone_init, missile_idx, n_fast=50000, n_refine=10):
    """快速+精确两阶段搜索最优起爆位置"""
    dx, dy, dz = drone_init
    v_min, v_max = cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX
    M1_vx = cfg.MISSILE_SPEED * abs(cfg.MISSILES_DIR[missile_idx][0])
    M1_start_x = cfg.MISSILES_INIT[missile_idx][0]

    best_val = 0
    best_params = None
    found = 0

    # 阶段1: 快速粗筛
    for i in range(n_fast):
        # 智能采样
        x = np.random.uniform(max(3000, dx-3000), dx+100)
        y = np.random.uniform(min(0, dy-300), max(dy+300, 250))
        z = np.random.uniform(max(100, dz-600), dz-5)

        drone_h = np.sqrt((x-dx)**2 + (y-dy)**2)
        if drone_h < 1 or z >= dz: continue

        fall_dist = dz - z
        td = np.sqrt(2*fall_dist/cfg.G)

        feasible = False
        best_config = None
        for v in [v_max, 115, 100, v_min]:
            tr = drone_h/v - td
            if tr >= 0:
                total_t = tr + td
                M1_arr = (M1_start_x - x)/M1_vx
                if M1_arr > 0 and total_t <= M1_arr:
                    theta = np.arctan2(y-dy, x-dx)
                    feasible = True
                    best_config = (theta, v, tr, td)
                    break
        if not feasible: continue

        theta, v, tr, td = best_config
        val = simulate_single_bomb(drone_init, theta, v, tr, td, missile_idx, KP_FAST, 0.02, 25.)

        if val > 0.01:
            found += 1
            if val > best_val:
                best_val = val
                best_params = (theta, v, tr, td, x, y, z)

    if best_params is None:
        return None, 0

    print(f"  粗筛找到{found}个, 最优≈{best_val:.1f}s", flush=True)

    # 阶段2: 精评top candidates
    # (简化为只精确评估最优的一个)
    theta, v, tr, td, x, y, z = best_params
    val_full = full_eval_1bomb(drone_init, theta, v, tr, td, missile_idx)
    print(f"  精评: {val_full:.4f}s", flush=True)

    return {'theta': theta, 'speed': v, 'release_time': tr, 'delay': td,
            'pos': (x, y, z)}, val_full


def main():
    print("="*70)
    print("  C题 烟幕干扰弹投放策略 - 最终求解")
    print("="*70)
    t0_total = time.time()

    # === P1 ===
    print("\n### 问题1 ###", flush=True)
    p1 = simulate_single_bomb(cfg.DRONES_INIT[0], np.pi, 120., 1.2, 3.2, 0,
                               KP_FULL, cfg.DT_FINE, 30.)
    print(f"有效遮蔽时长: {p1:.4f} s", flush=True)

    # === P2 ===
    print("\n### 问题2 ###", flush=True)
    t0 = time.time()
    params, p2 = search_best_detonation(cfg.DRONES_INIT[0], 0, n_fast=30000)
    t2 = time.time()-t0

    if params:
        d = np.array([np.cos(params['theta']), np.sin(params['theta']), 0.])
        di = cfg.DRONES_INIT[0]
        dp = di + params['speed']*d*(params['release_time']+params['delay'])
        dp[2] -= 0.5*9.8*params['delay']**2
        print(f"θ={np.degrees(params['theta']):.1f}° v={params['speed']:.0f}m/s "
              f"tr={params['release_time']:.4f}s td={params['delay']:.4f}s", flush=True)
        print(f"起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})", flush=True)
        print(f"最优: {p2:.4f}s ({t2:.0f}s)", flush=True)

    # === P3 ===
    print("\n### 问题3 ###", flush=True)
    t0 = time.time()
    if params:
        theta = params['theta']; speed = params['speed']
        td_ref = params['delay']

        # 快速搜索最佳三弹时序
        best_p3 = 0
        best_rt = None; best_dd = None
        for i in range(2000):
            rt1 = np.random.uniform(0, 5)
            int2 = np.random.uniform(1, 5); int3 = np.random.uniform(1, 5)
            rt = np.array([rt1, rt1+int2, rt1+int2+int3])
            dd = np.array([td_ref + np.random.uniform(-1, 1) for _ in range(3)])
            dd = np.clip(dd, 1, 8)
            val = fast_eval_3bombs(cfg.DRONES_INIT[0], theta, speed, rt, dd, np.array([0,0,0]))
            if val > best_p3:
                best_p3 = val; best_rt = rt.copy(); best_dd = dd.copy()

        if best_rt is not None:
            p3 = full_eval_3bombs(cfg.DRONES_INIT[0], theta, speed, best_rt, best_dd, np.array([0,0,0]))
            t3 = time.time()-t0
            for j in range(3):
                dp = cfg.DRONES_INIT[0] + speed*np.array([np.cos(theta), np.sin(theta), 0])*(best_rt[j]+best_dd[j])
                dp[2] -= 0.5*9.8*best_dd[j]**2
                print(f"  弹{j+1}: tr={best_rt[j]:.4f}s td={best_dd[j]:.4f}s 起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
            print(f"总遮蔽: {p3:.4f}s ({t3:.0f}s)", flush=True)
        else:
            p3 = 0
    else:
        p3 = 0

    # === P4 ===
    print("\n### 问题4 ###", flush=True)
    t0 = time.time()
    p4_results = []
    for idx in [0, 1, 2]:
        name = ['FY1','FY2','FY3'][idx]
        print(f"--- {name} ---", flush=True)
        p, v = search_best_detonation(cfg.DRONES_INIT[idx], 0, n_fast=20000)
        if p:
            p4_results.append({**p, 'name': name, 'drone_idx': idx})
            print(f"  {name}: {v:.4f}s", flush=True)
        else:
            print(f"  {name}: 未找到!", flush=True)

    if p4_results:
        dps = []
        for r in p4_results:
            dps.append({
                'drone_init': cfg.DRONES_INIT[r['drone_idx']],
                'theta': r['theta'], 'speed': r['speed'],
                'release_times': np.array([r['release_time']]),
                'detonation_delays': np.array([r['delay']]),
                'missile_indices': [0],
            })
        p4, per4 = simulate_multi_drone_multi_bomb(dps, 0.005, 35.)
        t4 = time.time()-t0
        print(f"协同遮蔽: {p4:.4f}s ({t4:.0f}s)", flush=True)
    else:
        p4 = 0

    # === P5 ===
    print("\n### 问题5 ###", flush=True)
    t0 = time.time()
    order = cfg.INTERCEPT_ORDER
    names = ['FY1','FY2','FY3','FY4','FY5']
    p5_results = []

    for di in range(5):
        print(f"--- {names[di]} ---", flush=True)
        od = order[names[di]]
        bomb_cfgs = []
        for bi, mi in enumerate(od):
            p, v = search_best_detonation(cfg.DRONES_INIT[di], mi, n_fast=10000)
            if p: bomb_cfgs.append({**p, 'missile_idx': mi})

        if bomb_cfgs:
            thetas = [b['theta'] for b in bomb_cfgs]
            speeds = [b['speed'] for b in bomb_cfgs]
            theta_m = np.median(thetas)
            speed_m = np.median(speeds)

            rt = np.array([b['release_time'] for b in bomb_cfgs])
            for j in range(1, len(rt)):
                if rt[j] < rt[j-1]+1: rt[j] = rt[j-1]+1
            dd = np.array([b['delay'] for b in bomb_cfgs])
            mi_arr = np.array([b['missile_idx'] for b in bomb_cfgs])

            p5_results.append({
                'name': names[di], 'drone_idx': di,
                'theta': theta_m, 'speed': speed_m,
                'release_times': rt, 'delays': dd, 'missile_indices': mi_arr,
            })
            print(f"  θ={np.degrees(theta_m):.0f}° v={speed_m:.0f}m/s", flush=True)

    if p5_results:
        dps = []
        for r in p5_results:
            dps.append({
                'drone_init': cfg.DRONES_INIT[r['drone_idx']],
                'theta': r['theta'], 'speed': r['speed'],
                'release_times': r['release_times'],
                'detonation_delays': r['delays'],
                'missile_indices': list(r['missile_indices']),
            })
        p5, per5 = simulate_multi_drone_multi_bomb(dps, 0.005, 40.)
        t5 = time.time()-t0
        print(f"总遮蔽: {p5:.4f}s (M1:{per5[0]:.2f} M2:{per5[1]:.2f} M3:{per5[2]:.2f}) ({t5:.0f}s)", flush=True)
    else:
        p5 = 0; per5 = np.zeros(3)

    # Save Excel files
    _save_excel(p3 if 'p3' in dir() else 0, best_rt if 'best_rt' in dir() else None,
                best_dd if 'best_dd' in dir() else None, p4_results if 'p4_results' in dir() else [],
                p5_results, p4 if 'p4' in dir() else 0, p5 if 'p5' in dir() else 0)

    # Summary
    total_t = time.time()-t0_total
    print("\n" + "="*70)
    print("                    最终结果汇总 (C题)")
    print("="*70)
    print(f"  问题1: {p1:.4f} s")
    print(f"  问题2: {p2:.4f} s")
    print(f"  问题3: {p3:.4f} s  -> result1.xlsx")
    print(f"  问题4: {p4:.4f} s  -> result2.xlsx")
    print(f"  问题5: {p5:.4f} s  -> result3.xlsx")
    print(f"  总耗时: {total_t:.0f}s ({total_t/60:.1f}min)")
    print("="*70)


def _save_excel(p3_val, p3_rt, p3_dd, p4_results, p5_results, p4_val, p5_val):
    import openpyxl

    # result1.xlsx
    if p3_rt is not None:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题3"
        h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","投放时间s","延时s",
             "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
        for c,hv in enumerate(h,1): ws.cell(1,c,hv)
        di = cfg.DRONES_INIT[0]
        for j in range(3):
            rp = di + 80*np.array([-1,0,0])*p3_rt[j]
            dp = rp + 80*np.array([-1,0,0])*p3_dd[j]; dp[2]-=0.5*9.8*p3_dd[j]**2
            rw=j+2
            ws.cell(rw,1,"FY1"); ws.cell(rw,2,round(np.pi,6)); ws.cell(rw,3,180.0)
            ws.cell(rw,4,80); ws.cell(rw,5,j+1)
            ws.cell(rw,6,round(p3_rt[j],4)); ws.cell(rw,7,round(p3_dd[j],4))
            ws.cell(rw,8,round(rp[0],2)); ws.cell(rw,9,round(rp[1],2)); ws.cell(rw,10,round(rp[2],2))
            ws.cell(rw,11,round(dp[0],2)); ws.cell(rw,12,round(dp[1],2)); ws.cell(rw,13,round(dp[2],2))
            if j==0: ws.cell(rw,14,round(p3_val,4))
        wb.save("result1.xlsx"); print("已保存 result1.xlsx")

    # result2.xlsx
    if p4_results:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题4"
        h = ["无人机","航向角rad","航向角°","速度m/s","投放时间s","延时s",
             "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
        for c,hv in enumerate(h,1): ws.cell(1,c,hv)
        for i,r in enumerate(p4_results):
            d = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.])
            di = cfg.DRONES_INIT[r['drone_idx']]
            rp = di + r['speed']*d*r['release_time']
            dp = rp + r['speed']*d*r['delay']; dp[2]-=0.5*9.8*r['delay']**2
            rw=i+2
            ws.cell(rw,1,r['name']); ws.cell(rw,2,round(r['theta'],6))
            ws.cell(rw,3,round(np.degrees(r['theta']),4)); ws.cell(rw,4,round(r['speed'],2))
            ws.cell(rw,5,round(r['release_time'],4)); ws.cell(rw,6,round(r['delay'],4))
            ws.cell(rw,7,round(rp[0],2)); ws.cell(rw,8,round(rp[1],2)); ws.cell(rw,9,round(rp[2],2))
            ws.cell(rw,10,round(dp[0],2)); ws.cell(rw,11,round(dp[1],2)); ws.cell(rw,12,round(dp[2],2))
            if i==0: ws.cell(rw,13,round(p4_val,4))
        wb.save("result2.xlsx"); print("已保存 result2.xlsx")

    # result3.xlsx
    if p5_results:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题5"
        h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","目标导弹","投放时间s","延时s",
             "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
        for c,hv in enumerate(h,1): ws.cell(1,c,hv)
        rw=2
        for r in p5_results:
            d = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.])
            di = cfg.DRONES_INIT[r['drone_idx']]
            for j in range(len(r['release_times'])):
                rp = di + r['speed']*d*r['release_times'][j]
                dp = rp + r['speed']*d*r['delays'][j]; dp[2]-=0.5*9.8*r['delays'][j]**2
                ws.cell(rw,1,r['name']); ws.cell(rw,2,round(r['theta'],6))
                ws.cell(rw,3,round(np.degrees(r['theta']),4)); ws.cell(rw,4,round(r['speed'],2))
                ws.cell(rw,5,j+1); ws.cell(rw,6,f"M{int(r['missile_indices'][j])+1}")
                ws.cell(rw,7,round(r['release_times'][j],4)); ws.cell(rw,8,round(r['delays'][j],4))
                ws.cell(rw,9,round(rp[0],2)); ws.cell(rw,10,round(rp[1],2)); ws.cell(rw,11,round(rp[2],2))
                ws.cell(rw,12,round(dp[0],2)); ws.cell(rw,13,round(dp[1],2)); ws.cell(rw,14,round(dp[2],2))
                if rw==2: ws.cell(rw,15,round(p5_val,4))
                rw+=1
        wb.save("result3.xlsx"); print("已保存 result3.xlsx")


if __name__ == "__main__":
    main()
