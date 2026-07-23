"""
问题3: 利用无人机FY1投放3枚烟幕干扰弹实施对M1的干扰
输出结果到 result1.xlsx

优化变量(8维): [theta, speed, release1, interval2, interval3, delay1, delay2, delay3]
实际投放时间: [release1, release1+interval2, release1+interval2+interval3]
"""
import numpy as np
import openpyxl
from config import (
    DRONES_INIT, DRONE_SPEED_MIN, DRONE_SPEED_MAX,
    BOMB_INTERVAL_MIN, DT, T_TOTAL,
)
from simulation import simulate_multi_bomb_single_drone, get_target_keypoints
from pso import PSO, local_polish
from problem2 import Problem2Objective

# 为 problem3 使用独立的 PSO 参数
PSO_SWARM_SIZE_P3 = 300
PSO_MAX_ITER_P3 = 150

# 阶段0(热启动预搜)的PSO参数，问题规模小(单弹4维)，预算给小一些即可
PSO_SWARM_SIZE_P3_SEED = 100
PSO_MAX_ITER_P3_SEED = 40


class Problem3Objective:
    """模块级可pickle的目标函数对象，供PSO多进程worker调用"""

    def __init__(self, drone_init, target_keypoints):
        self.drone_init = drone_init
        self.target_keypoints = target_keypoints

    def __call__(self, x):
        theta = x[0]
        speed = x[1]
        release_times = np.array([
            x[2],
            x[2] + x[3],
            x[2] + x[3] + x[4],
        ])
        detonation_delays = np.array([x[5], x[6], x[7]])
        missile_indices = np.array([0, 0, 0])  # 全部针对M1

        return simulate_multi_bomb_single_drone(
            self.drone_init, theta, speed, release_times, detonation_delays,
            missile_indices, target_keypoints=self.target_keypoints, dt=DT, t_total=T_TOTAL
        )


def solve_problem3():
    """求解问题3: 单机三弹最优策略"""
    print("=" * 60)
    print("问题3: FY1投放3枚烟幕干扰弹对M1的最优策略 (PSO优化)")
    print("=" * 60)

    drone_init = DRONES_INIT[0]  # FY1
    target_keypoints = get_target_keypoints(360, 10)

    # 决策变量(8维):
    # x[0] = theta, x[1] = speed
    # x[2] = release1, x[3] = interval_to_2nd, x[4] = interval_to_3rd
    # x[5] = delay1, x[6] = delay2, x[7] = delay3
    bounds = [
        (2.73, 3.53),     # theta (rad) - 朝向真目标方向附近(原范围方向错误，见problem2.py)
        (DRONE_SPEED_MIN, DRONE_SPEED_MAX),  # speed
        (0.0, 5.0),                      # release1
        (BOMB_INTERVAL_MIN, 4.0),        # interval 1->2
        (BOMB_INTERVAL_MIN, 4.0),        # interval 2->3
        (0.0, 6.0),                      # delay1
        (0.0, 6.0),                      # delay2
        (0.0, 6.0),                      # delay3
    ]

    # ============================================================
    # 阶段0: 用问题2同款的单弹目标函数快速预搜一个"够用"的起点，
    # 给下面8维PSO当热启动种子——对应国奖论文里"贪心算法找可接受解，
    # PSO在其附近精修"的思路，而不是让PSO从纯随机初始化的8维空间里摸索。
    # ============================================================
    print("\n阶段0: 单弹快速预搜(为8维PSO提供热启动起点)...")
    seed_kp = get_target_keypoints(360, 0)  # 用问题2同档精度(无侧面点)，预搜更快
    seed_obj = Problem2Objective(drone_init, seed_kp)
    seed_bounds = [bounds[0], bounds[1], (0.0, 5.0), (0.0, 6.0)]
    seed_pso = PSO(seed_obj, seed_bounds, n_particles=PSO_SWARM_SIZE_P3_SEED,
                   max_iter=PSO_MAX_ITER_P3_SEED, maximize=True, verbose=False)
    seed_x, seed_f = seed_pso.optimize()
    print(f"  预搜起点: θ={np.degrees(seed_x[0]):.1f}° v={seed_x[1]:.1f}m/s "
          f"release={seed_x[2]:.2f}s delay={seed_x[3]:.2f}s (单弹遮蔽{seed_f:.4f}s)")

    # 用这个单弹解构造8维种子: 三发弹依次按最小间隔错开投放，起爆延时先沿用预搜结果，
    # 后续PSO会在这个起点附近继续搜索(种子只替换初始种群里的一个粒子，不锁死解)
    seed_moderate_delay = [seed_x[0], seed_x[1], seed_x[2], BOMB_INTERVAL_MIN, BOMB_INTERVAL_MIN,
                            seed_x[3], seed_x[3], seed_x[3]]

    # 第二个种子: 参考学长MATLAB代码(m3.m)里手动收窄过的搜索范围——"贴近速度上限+
    # 第一发几乎零延时"这个策略分支。用A题参数验证过: PSO单靠上面那个(低速+中等延时)
    # 种子只能收敛到5.51s，加上这个种子之后第40代就到7.64s，逼近参考值7.65s。
    # 速度/时序都按当前config的实际范围换算，不写死A题验证时的绝对数值，换成C题范围
    # 也能用。
    speed_hi = DRONE_SPEED_MAX - 0.1 * (DRONE_SPEED_MAX - DRONE_SPEED_MIN)
    seed_high_speed_low_delay = [
        seed_x[0],           # 复用同一个热启动搜到的方向
        speed_hi,             # 速度贴近上限
        0.1,                  # release1: 尽早投放
        BOMB_INTERVAL_MIN,    # interval2: 最小间隔
        BOMB_INTERVAL_MIN,    # interval3: 最小间隔
        0.1,                  # delay1: 几乎零延时
        3.0,                  # delay2
        3.0,                  # delay3
    ]

    seed_positions = [seed_moderate_delay, seed_high_speed_low_delay]

    objective = Problem3Objective(drone_init, target_keypoints)

    print(f"\n变量维度: 8 (theta, speed, 3×release, 3×delay)")
    print(f"粒子群规模: {PSO_SWARM_SIZE_P3}, 迭代次数: {PSO_MAX_ITER_P3}")

    pso = PSO(objective, bounds, n_particles=PSO_SWARM_SIZE_P3,
              max_iter=PSO_MAX_ITER_P3, maximize=True, verbose=True,
              seed_positions=seed_positions)
    x_opt, f_opt = pso.optimize()

    # PSO收敛完之后做一次局部精修(对应论文里PSO+TS、MATLAB fmincon hybrid那一层收尾)，
    # 精修不保证一定更好(无梯度法不单调)，跟PSO原结果比较取较大的那个
    print("\n局部精修(Powell)...")
    x_polished, f_polished = local_polish(objective, x_opt, bounds)
    if f_polished > f_opt:
        print(f"  精修有提升: {f_opt:.4f}s -> {f_polished:.4f}s")
        x_opt, f_opt = x_polished, f_polished
    else:
        print(f"  精修没有提升({f_polished:.4f}s <= {f_opt:.4f}s)，保留PSO原结果")

    # 解析结果
    theta = x_opt[0]
    speed = x_opt[1]
    release_times = np.array([x_opt[2], x_opt[2] + x_opt[3], x_opt[2] + x_opt[3] + x_opt[4]])
    detonation_delays = np.array([x_opt[5], x_opt[6], x_opt[7]])

    direction = np.array([np.cos(theta), np.sin(theta), 0.0])

    print(f"\n优化结果:")
    print(f"  航向角θ: {theta:.4f} rad ({np.degrees(theta):.2f}°)")
    print(f"  飞行速度: {speed:.2f} m/s")

    for i in range(3):
        release_pos = drone_init + speed * direction * release_times[i]
        detonation_pos = release_pos + speed * direction * detonation_delays[i]
        detonation_pos[2] -= 0.5 * 9.8 * detonation_delays[i] ** 2

        print(f"\n  烟幕弹{i+1}:")
        print(f"    投放时间: {release_times[i]:.4f} s")
        print(f"    起爆延时: {detonation_delays[i]:.4f} s")
        print(f"    投放点: ({release_pos[0]:.2f}, {release_pos[1]:.2f}, {release_pos[2]:.2f})")
        print(f"    起爆点: ({detonation_pos[0]:.2f}, {detonation_pos[1]:.2f}, {detonation_pos[2]:.2f})")

    print(f"\n  总有效遮蔽时长: {f_opt:.4f} s")

    # 保存到 result1.xlsx
    save_result1(theta, speed, release_times, detonation_delays, drone_init,
                 direction, f_opt)

    return x_opt, f_opt


def save_result1(theta, speed, release_times, detonation_delays,
                 drone_init, direction, total_time):
    """保存问题3结果到 result1.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问题3结果"

    # 表头
    headers = ["无人机", "航向角(rad)", "航向角(°)", "飞行速度(m/s)",
               "烟幕弹编号", "投放时间(s)", "起爆延时(s)",
               "投放点X", "投放点Y", "投放点Z",
               "起爆点X", "起爆点Y", "起爆点Z",
               "总有效遮蔽时长(s)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 数据行
    for i in range(3):
        row = i + 2
        release_pos = drone_init + speed * direction * release_times[i]
        detonation_pos = release_pos + speed * direction * detonation_delays[i]
        detonation_pos[2] -= 0.5 * 9.8 * detonation_delays[i] ** 2

        ws.cell(row=row, column=1, value="FY1")
        ws.cell(row=row, column=2, value=round(theta, 6))
        ws.cell(row=row, column=3, value=round(np.degrees(theta), 4))
        ws.cell(row=row, column=4, value=round(speed, 2))
        ws.cell(row=row, column=5, value=i + 1)
        ws.cell(row=row, column=6, value=round(release_times[i], 4))
        ws.cell(row=row, column=7, value=round(detonation_delays[i], 4))
        ws.cell(row=row, column=8, value=round(release_pos[0], 2))
        ws.cell(row=row, column=9, value=round(release_pos[1], 2))
        ws.cell(row=row, column=10, value=round(release_pos[2], 2))
        ws.cell(row=row, column=11, value=round(detonation_pos[0], 2))
        ws.cell(row=row, column=12, value=round(detonation_pos[1], 2))
        ws.cell(row=row, column=13, value=round(detonation_pos[2], 2))
        if i == 0:
            ws.cell(row=row, column=14, value=round(total_time, 4))

    filepath = "result1.xlsx"
    wb.save(filepath)
    print(f"\n结果已保存至: {filepath}")


if __name__ == "__main__":
    solve_problem3()
