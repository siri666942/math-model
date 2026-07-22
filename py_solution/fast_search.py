"""
快速几何搜索 - 利用几何约束大幅缩小搜索空间
"""
import numpy as np
import config as cfg
from simulation import simulate_single_bomb, get_target_keypoints


def fast_find_best(drone_init, missile_idx=0, n_samples=10000):
    """
    快速找到最优起爆位置

    策略:
    1. 利用几何约束过滤绝大多数不可行位置
    2. 在可行区域内高效采样
    3. 两阶段评估: 粗筛→精评
    """
    kp_fast = get_target_keypoints(18, 0)   # 36点用于粗筛
    kp_full = get_target_keypoints(360, 0)   # 720点用于精评

    M1_init = cfg.MISSILES_INIT[missile_idx]
    M1_dir = cfg.MISSILES_DIR[missile_idx]
    M1_vx = cfg.MISSILE_SPEED * abs(M1_dir[0])  # M1 x方向速度
    M1_start_x = M1_init[0]

    dx, dy, dz = drone_init
    v_min, v_max = cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX

    best_val = 0
    best_params = None

    print(f"  采样 {n_samples} 个候选...")
    found = 0

    for i in range(n_samples):
        if (i + 1) % 5000 == 0:
            print(f"    [{i+1}/{n_samples}] 找到: {found}, 最优: {best_val:.4f}s")

        # === 智能采样: 限制在几何可行区域 ===
        # 1. x: 目标→M1之间，且无人机能在导弹前到达
        #    最小x = drone能到达的最远点 (在最大速度下)
        drone_to_target = np.sqrt((dx)**2 + (dy - 200)**2)
        max_reach_time = drone_to_target / v_min
        min_x = dx - v_max * 15  # 15s内能到达

        # 更实际的约束:
        # 从导弹到达x的时间: t_M1(x) = (M1_start_x - x) / M1_vx
        # 无人机到达(x,y)的最短时间: t_drone(x,y) = sqrt((x-dx)²+(y-dy)²) / v_max
        # 需要: t_drone < t_M1

        # 起爆位置应该靠近无人机初始位置(因为要在导弹到达前起爆)
        # 重点采样靠近无人机的区域
        # x: 无人机初始x到无人机初始x-2000之间
        x = np.random.uniform(max(3000, dx - 3000), dx + 100)

        # y: 目标在y=200, M1在y=0
        # 起爆y应在0~200之间(导弹到目标的y范围内)
        y_min = min(dy - 300, 0)
        y_max = max(dy + 300, 250)
        y = np.random.uniform(y_min, y_max)

        # 快速可行性: 无人机到达时间 vs 导弹到达时间
        drone_time_to_xy = np.sqrt((x-dx)**2 + (y-dy)**2) / v_max
        missile_time_to_x = (M1_start_x - x) / M1_vx

        if missile_time_to_x <= 0 or drone_time_to_xy >= missile_time_to_x:
            continue  # 无人机无法在导弹之前到达

        # z: 烟幕弹从无人机高度自由落体
        # 起爆时有 td = sqrt(2*(dz-z)/g)
        # 总时间 = tr + td = drone_time + td
        # 需要总时间 < 导弹到达时间
        # 尝试在合理范围内采样z
        # z: 起爆高度应在无人机高度以下，但不太低
        # 自由落体时间 td = sqrt(2*(dz-z)/g)
        # td 太大则总时间过大，太小则tr可能为负
        z = np.random.uniform(max(100, dz - 600), dz - 5)

        # 详细可行性检查
        drone_to_pt_h = np.sqrt((x-dx)**2 + (y-dy)**2)
        fall_dist = dz - z
        if fall_dist <= 0:
            continue

        td = np.sqrt(2 * fall_dist / cfg.G)

        # 试不同速度
        feasible = False
        best_config = None
        for v_test in [v_max, 115, 100, v_min]:
            tr = drone_to_pt_h / v_test - td
            if tr < 0:
                continue
            total_time = tr + td
            missile_arrival = (M1_start_x - x) / M1_vx
            if missile_arrival > 0 and total_time <= missile_arrival:
                theta = np.arctan2(y - dy, x - dx)
                feasible = True
                best_config = (theta, v_test, tr, td)
                break

        if not feasible:
            continue

        theta, v_test, tr, td = best_config

        # 粗筛
        val = simulate_single_bomb(drone_init, theta, v_test, tr, td,
                                    missile_idx, kp_fast, 0.01, 25.0)

        if val > 0.01:
            found += 1
            # 精评
            val_full = simulate_single_bomb(drone_init, theta, v_test, tr, td,
                                             missile_idx, kp_full, 0.005, 30.0)
            if val_full > best_val:
                best_val = val_full
                best_params = {
                    'pos': (x, y, z), 'theta': theta, 'speed': v_test,
                    'release_time': tr, 'delay': td,
                }
                print(f"    *** 新最优: {best_val:.4f}s "
                      f"({x:.0f},{y:.0f},{z:.0f}) "
                      f"θ={np.degrees(theta):.1f}° v={v_test:.0f}")

    return best_params, best_val


def run_all_problems_fast():
    """快速运行所有问题"""
    import time
    results = {}

    # P1: 固定参数
    print("="*60)
    print("问题1: FY1 -> M1")
    print("="*60)
    t0 = time.time()
    p1 = simulate_single_bomb(cfg.DRONES_INIT[0], np.pi, 120., 1.2, 3.2, 0,
                               dt=cfg.DT_FINE, t_total=30.)
    results['p1'] = p1
    print(f"结果: {p1:.4f}s ({time.time()-t0:.1f}s)")

    # P2: FY1 + 1 bomb -> M1
    print("\n" + "="*60)
    print("问题2: FY1 + 1弹 -> M1")
    print("="*60)
    t0 = time.time()
    params, val = fast_find_best(cfg.DRONES_INIT[0], 0, n_samples=20000)
    results['p2'] = (params, val)
    if params:
        d = np.array([np.cos(params['theta']), np.sin(params['theta']), 0.])
        di = cfg.DRONES_INIT[0]
        rp = di + params['speed']*d*params['release_time']
        dp = rp + params['speed']*d*params['delay']; dp[2] -= 0.5*9.8*params['delay']**2
        print(f"θ={np.degrees(params['theta']):.1f}° v={params['speed']:.0f}m/s "
              f"tr={params['release_time']:.4f}s td={params['delay']:.4f}s")
        print(f"起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
        print(f"最优遮蔽: {val:.4f}s ({time.time()-t0:.1f}s)")
    else:
        print("未找到!")

    # P3: FY1 + 3 bombs -> M1 (simplified: build on P2 result)
    print("\n" + "="*60)
    print("问题3: FY1 + 3弹 -> M1 (基于P2扩展)")
    print("="*60)
    t0 = time.time()
    if params:
        # 使用P2的最优参数，在时间上错开3枚弹
        theta = params['theta']
        speed = params['speed']
        # 三枚弹: 间隔至少1s，在P2最优投放时间附近分布
        rt0 = max(0, params['release_time'] - 2)
        rt = np.array([rt0, rt0 + 1.5, rt0 + 3.0])
        dd = np.array([params['delay'], params['delay'], params['delay']])

        from simulation import simulate_multi_bomb_single_drone
        val3 = simulate_multi_bomb_single_drone(
            cfg.DRONES_INIT[0], theta, speed, rt, dd,
            np.array([0,0,0]), get_target_keypoints(360,10), 0.005, 30.)
        results['p3'] = (theta, speed, rt, dd, val3)
        print(f"θ={np.degrees(theta):.1f}° v={speed:.0f}m/s")
        for j in range(3):
            print(f"  弹{j+1}: 投放={rt[j]:.4f}s 延时={dd[j]:.4f}s")
        print(f"总遮蔽: {val3:.4f}s ({time.time()-t0:.1f}s)")
    else:
        results['p3'] = (0, 100, np.array([0,1,2]), np.array([3,3,3]), 0)
        print("使用默认参数")

    # P4: FY1+FY2+FY3 -> M1
    print("\n" + "="*60)
    print("问题4: FY1/FY2/FY3 -> M1")
    print("="*60)
    t0 = time.time()
    p4_results = []
    for idx in [0, 1, 2]:
        name = ['FY1','FY2','FY3'][idx]
        print(f"\n--- {name} ---")
        p, v = fast_find_best(cfg.DRONES_INIT[idx], 0, n_samples=15000)
        if p:
            p4_results.append({**p, 'name': name, 'drone_idx': idx})
            print(f"  {name}: {v:.4f}s")

    if p4_results:
        from simulation import simulate_multi_drone_multi_bomb
        dps = []
        for r in p4_results:
            dps.append({
                'drone_init': cfg.DRONES_INIT[r['drone_idx']],
                'theta': r['theta'], 'speed': r['speed'],
                'release_times': np.array([r['release_time']]),
                'detonation_delays': np.array([r['delay']]),
                'missile_indices': [0],
            })
        total4, per4 = simulate_multi_drone_multi_bomb(dps, 0.005, 35.)
        results['p4'] = (p4_results, total4)
        print(f"\n协同遮蔽: {total4:.4f}s ({time.time()-t0:.1f}s)")
    else:
        results['p4'] = ([], 0)

    # P5: 5 drones -> M1/M2/M3
    print("\n" + "="*60)
    print("问题5: 5机协同 -> M1/M2/M3")
    print("="*60)
    t0 = time.time()
    p5_results = []
    order = cfg.INTERCEPT_ORDER
    names = ['FY1','FY2','FY3','FY4','FY5']

    for di in range(5):
        print(f"\n--- {names[di]} ---")
        od = order[names[di]]
        bomb_configs = []
        for bi, mi in enumerate(od):
            p, v = fast_find_best(cfg.DRONES_INIT[di], mi, n_samples=8000)
            if p:
                bomb_configs.append({**p, 'missile_idx': mi})

        if bomb_configs:
            # 统一航向速度(取各弹结果的中位数)
            theta_vals = [b['theta'] for b in bomb_configs]
            theta = np.median(theta_vals)
            speed_vals = [b['speed'] for b in bomb_configs]
            speed = np.median(speed_vals)

            rt = np.array([b['release_time'] for b in bomb_configs])
            # 确保间隔>=1s
            for j in range(1, len(rt)):
                if rt[j] < rt[j-1] + 1.0:
                    rt[j] = rt[j-1] + 1.0
            dd = np.array([b['delay'] for b in bomb_configs])
            mi = np.array([b['missile_idx'] for b in bomb_configs])

            p5_results.append({
                'name': names[di], 'drone_idx': di,
                'theta': theta, 'speed': speed,
                'release_times': rt, 'delays': dd,
                'missile_indices': mi,
            })
            print(f"  θ={np.degrees(theta):.1f}° v={speed:.0f}m/s")

    if p5_results:
        from simulation import simulate_multi_drone_multi_bomb
        dps = []
        for r in p5_results:
            dps.append({
                'drone_init': cfg.DRONES_INIT[r['drone_idx']],
                'theta': r['theta'], 'speed': r['speed'],
                'release_times': r['release_times'],
                'detonation_delays': r['delays'],
                'missile_indices': list(r['missile_indices']),
            })
        total5, per5 = simulate_multi_drone_multi_bomb(dps, 0.005, 40.)
        results['p5'] = (p5_results, total5, per5)
        print(f"\n总遮蔽: {total5:.4f}s (M1:{per5[0]:.2f} M2:{per5[1]:.2f} M3:{per5[2]:.2f}) ({time.time()-t0:.1f}s)")

    # Save Excel files
    _save_results(results)

    return results


def _save_results(results):
    """保存结果到Excel"""
    import openpyxl

    # result1.xlsx (P3)
    if 'p3' in results and results['p3'][4] > 0:
        theta, speed, rt, dd, val = results['p3']
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题3"
        h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","投放时间s","延时s",
             "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
        for c,hv in enumerate(h,1): ws.cell(1,c,hv)
        d = np.array([np.cos(theta), np.sin(theta), 0.])
        di = cfg.DRONES_INIT[0]
        for j in range(3):
            rp = di + speed*d*rt[j]
            dp = rp + speed*d*dd[j]; dp[2] -= 0.5*9.8*dd[j]**2
            rw = j+2
            ws.cell(rw,1,"FY1"); ws.cell(rw,2,round(theta,6)); ws.cell(rw,3,round(np.degrees(theta),4))
            ws.cell(rw,4,round(speed,2)); ws.cell(rw,5,j+1)
            ws.cell(rw,6,round(rt[j],4)); ws.cell(rw,7,round(dd[j],4))
            ws.cell(rw,8,round(rp[0],2)); ws.cell(rw,9,round(rp[1],2)); ws.cell(rw,10,round(rp[2],2))
            ws.cell(rw,11,round(dp[0],2)); ws.cell(rw,12,round(dp[1],2)); ws.cell(rw,13,round(dp[2],2))
            if j==0: ws.cell(rw,14,round(val,4))
        wb.save("result1.xlsx"); print("已保存 result1.xlsx")

    # result2.xlsx (P4)
    if 'p4' in results and results['p4'][0]:
        p4r, total = results['p4']
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题4"
        h = ["无人机","航向角rad","航向角°","速度m/s","投放时间s","延时s",
             "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
        for c,hv in enumerate(h,1): ws.cell(1,c,hv)
        for i,r in enumerate(p4r):
            d = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.])
            di = cfg.DRONES_INIT[r['drone_idx']]
            rp = di + r['speed']*d*r['release_time']
            dp = rp + r['speed']*d*r['delay']; dp[2] -= 0.5*9.8*r['delay']**2
            rw = i+2
            ws.cell(rw,1,r['name']); ws.cell(rw,2,round(r['theta'],6)); ws.cell(rw,3,round(np.degrees(r['theta']),4))
            ws.cell(rw,4,round(r['speed'],2)); ws.cell(rw,5,round(r['release_time'],4)); ws.cell(rw,6,round(r['delay'],4))
            ws.cell(rw,7,round(rp[0],2)); ws.cell(rw,8,round(rp[1],2)); ws.cell(rw,9,round(rp[2],2))
            ws.cell(rw,10,round(dp[0],2)); ws.cell(rw,11,round(dp[1],2)); ws.cell(rw,12,round(dp[2],2))
            if i==0: ws.cell(rw,13,round(total,4))
        wb.save("result2.xlsx"); print("已保存 result2.xlsx")

    # result3.xlsx (P5)
    if 'p5' in results and results['p5'][0]:
        p5r, total, per = results['p5']
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题5"
        h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","目标导弹","投放时间s","延时s",
             "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
        for c,hv in enumerate(h,1): ws.cell(1,c,hv)
        rw = 2
        for r in p5r:
            d = np.array([np.cos(r['theta']), np.sin(r['theta']), 0.])
            di = cfg.DRONES_INIT[r['drone_idx']]
            for j in range(len(r['release_times'])):
                rp = di + r['speed']*d*r['release_times'][j]
                dp = rp + r['speed']*d*r['delays'][j]; dp[2] -= 0.5*9.8*r['delays'][j]**2
                ws.cell(rw,1,r['name']); ws.cell(rw,2,round(r['theta'],6))
                ws.cell(rw,3,round(np.degrees(r['theta']),4)); ws.cell(rw,4,round(r['speed'],2))
                ws.cell(rw,5,j+1); ws.cell(rw,6,f"M{int(r['missile_indices'][j])+1}")
                ws.cell(rw,7,round(r['release_times'][j],4)); ws.cell(rw,8,round(r['delays'][j],4))
                ws.cell(rw,9,round(rp[0],2)); ws.cell(rw,10,round(rp[1],2)); ws.cell(rw,11,round(rp[2],2))
                ws.cell(rw,12,round(dp[0],2)); ws.cell(rw,13,round(dp[1],2)); ws.cell(rw,14,round(dp[2],2))
                if rw==2: ws.cell(rw,15,round(total,4))
                rw += 1
        wb.save("result3.xlsx"); print("已保存 result3.xlsx")


if __name__ == "__main__":
    import time
    t0 = time.time()
    results = run_all_problems_fast()
    elapsed = time.time() - t0

    print("\n" + "="*70)
    print("                    最终结果汇总 (C题)")
    print("="*70)
    p1 = results.get('p1', 0)
    p2 = results['p2'][1] if results.get('p2') and results['p2'][0] else 0
    p3 = results['p3'][4] if results.get('p3') else 0
    p4 = results['p4'][1] if results.get('p4') and results['p4'][0] else 0
    p5 = results['p5'][1] if results.get('p5') and results['p5'][0] else 0
    print(f"  问题1: {p1:.4f} s")
    print(f"  问题2: {p2:.4f} s")
    print(f"  问题3: {p3:.4f} s")
    print(f"  问题4: {p4:.4f} s")
    print(f"  问题5: {p5:.4f} s")
    print(f"  总耗时: {elapsed:.0f}s ({elapsed/60:.1f}min)")
    print("="*70)
