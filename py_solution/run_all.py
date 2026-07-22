"""
运行所有5个问题 - 使用C题参数
"""
import sys, os, time
import numpy as np

os.chdir(os.path.dirname(os.path.abspath(__file__)))

import config as cfg
from simulation import (
    simulate_single_bomb, simulate_multi_bomb_single_drone,
    simulate_multi_drone_multi_bomb, get_target_keypoints
)
from pso import PSO


def run_problem1():
    """问题1: 固定参数验证"""
    print("\n" + "="*60)
    print("问题1: FY1投放1枚烟幕弹对M1 - 给定参数计算")
    print("="*60)

    drone_init = cfg.DRONES_INIT[0]
    t = simulate_single_bomb(
        drone_init, cfg.P1_DRONE_THETA, cfg.P1_DRONE_SPEED,
        cfg.P1_RELEASE_TIME, cfg.P1_DETONATION_DELAY,
        missile_idx=0, dt=cfg.DT_FINE, t_total=30.0
    )
    print(f"参数: θ=π(180°), v=120m/s, 投放t={cfg.P1_RELEASE_TIME}s, 延时={cfg.P1_DETONATION_DELAY}s")
    print(f"结果: 有效遮蔽时长 = {t:.4f} s")
    return t


def run_problem2():
    """问题2: 单机单弹优化"""
    print("\n" + "="*60)
    print("问题2: FY1单机单弹最优策略 (PSO)")
    print("="*60)

    di = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(720, 0)

    bounds = [
        (0.7, 1.6),             # theta (~0.22π to ~0.51π)
        (cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX),
        (0.0, 12.0),
        (0.0, 8.0),
    ]

    def obj(x):
        return simulate_single_bomb(di, x[0], x[1], x[2], x[3], 0, kp, cfg.DT, 30.0)

    pso = PSO(obj, bounds, n_particles=300, max_iter=200, maximize=True, verbose=True)
    x, f = pso.optimize()

    d = np.array([np.cos(x[0]), np.sin(x[0]), 0.])
    rp = di + x[1]*d*x[2]
    dp = rp + x[1]*d*x[3]; dp[2] -= 0.5*9.8*x[3]**2

    print(f"\nθ={x[0]:.4f}rad({np.degrees(x[0]):.1f}°), v={x[1]:.1f}m/s")
    print(f"投放t={x[2]:.4f}s, 延时={x[3]:.4f}s")
    print(f"投放点:({rp[0]:.0f},{rp[1]:.0f},{rp[2]:.0f})")
    print(f"起爆点:({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
    print(f"最优遮蔽时长: {f:.4f} s")
    return x, f


def run_problem3():
    """问题3: 单机三弹 -> result1.xlsx"""
    print("\n" + "="*60)
    print("问题3: FY1三弹最优策略 (PSO)")
    print("="*60)

    di = cfg.DRONES_INIT[0]
    kp = get_target_keypoints(360, 5)

    bounds = [
        (0.7, 1.6),
        (cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX),
        (0.0, 5.0), (cfg.BOMB_INTERVAL_MIN, 5.0), (cfg.BOMB_INTERVAL_MIN, 5.0),
        (0.0, 8.0), (0.0, 8.0), (0.0, 8.0),
    ]

    def obj(x):
        rt = np.array([x[2], x[2]+x[3], x[2]+x[3]+x[4]])
        dd = np.array([x[5], x[6], x[7]])
        return simulate_multi_bomb_single_drone(di, x[0], x[1], rt, dd,
                                                 np.array([0,0,0]), kp, cfg.DT, 25.0)

    pso = PSO(obj, bounds, n_particles=300, max_iter=200, maximize=True, verbose=True)
    x, f = pso.optimize()

    rt = np.array([x[2], x[2]+x[3], x[2]+x[3]+x[4]])
    dd = np.array([x[5], x[6], x[7]])
    d = np.array([np.cos(x[0]), np.sin(x[0]), 0.])

    print(f"\nθ={x[0]:.4f}rad({np.degrees(x[0]):.1f}°), v={x[1]:.1f}m/s")
    for i in range(3):
        rp = di + x[1]*d*rt[i]
        dp = rp + x[1]*d*dd[i]; dp[2] -= 0.5*9.8*dd[i]**2
        print(f"  弹{i+1}: 投放t={rt[i]:.4f}s, 延时={dd[i]:.4f}s, 起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
    print(f"总遮蔽时长: {f:.4f} s")

    # Save result1.xlsx
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "问题3"
    h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","投放时间s","延时s",
         "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
    for c,hv in enumerate(h,1): ws.cell(1,c,hv)
    for i in range(3):
        rp = di + x[1]*d*rt[i]; dp = rp + x[1]*d*dd[i]; dp[2] -= 0.5*9.8*dd[i]**2
        rw = i+2
        ws.cell(rw,1,"FY1"); ws.cell(rw,2,round(x[0],6)); ws.cell(rw,3,round(np.degrees(x[0]),4))
        ws.cell(rw,4,round(x[1],2)); ws.cell(rw,5,i+1)
        ws.cell(rw,6,round(rt[i],4)); ws.cell(rw,7,round(dd[i],4))
        ws.cell(rw,8,round(rp[0],2)); ws.cell(rw,9,round(rp[1],2)); ws.cell(rw,10,round(rp[2],2))
        ws.cell(rw,11,round(dp[0],2)); ws.cell(rw,12,round(dp[1],2)); ws.cell(rw,13,round(dp[2],2))
        if i==0: ws.cell(rw,14,round(f,4))
    wb.save("result1.xlsx")
    print("已保存 result1.xlsx")
    return x, f


def run_problem4():
    """问题4: 三机各一弹 -> result2.xlsx"""
    print("\n" + "="*60)
    print("问题4: FY1/FY2/FY3协同 (PSO)")
    print("="*60)

    bounds = []
    # theta for 3 drones
    bounds += [(0.7, 1.6), (0.0, 0.8), (-0.8, 0.0)]
    # speed
    bounds += [(cfg.DRONE_SPEED_MIN, cfg.DRONE_SPEED_MAX)] * 3
    # release_time
    bounds += [(0.0, 20.0)] * 3
    # delay
    bounds += [(0.0, 20.0)] * 3

    def obj(x):
        th = x[0:3]; sp = x[3:6]; rt = x[6:9]; dl = x[9:12]
        dps = []
        for i in range(3):
            dps.append({
                'drone_init': cfg.DRONES_INIT[i], 'theta': th[i], 'speed': sp[i],
                'release_times': np.array([rt[i]]), 'detonation_delays': np.array([dl[i]]),
                'missile_indices': [0],
            })
        tt, _ = simulate_multi_drone_multi_bomb(dps, cfg.DT, 30.0)
        return tt

    pso = PSO(obj, bounds, n_particles=400, max_iter=200, maximize=True, verbose=True)
    x, f = pso.optimize()

    th = x[0:3]; sp = x[3:6]; rt = x[6:9]; dl = x[9:12]
    names = ['FY1','FY2','FY3']
    print(f"\n总遮蔽时长: {f:.4f} s")
    for i in range(3):
        d = np.array([np.cos(th[i]), np.sin(th[i]), 0.])
        rp = cfg.DRONES_INIT[i] + sp[i]*d*rt[i]
        dp = rp + sp[i]*d*dl[i]; dp[2] -= 0.5*9.8*dl[i]**2
        print(f"  {names[i]}: θ={np.degrees(th[i]):.1f}°, v={sp[i]:.1f}m/s, 投放t={rt[i]:.4f}s, 延时={dl[i]:.4f}s, 起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "问题4"
    h = ["无人机","航向角rad","航向角°","速度m/s","投放时间s","延时s",
         "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
    for c,hv in enumerate(h,1): ws.cell(1,c,hv)
    for i in range(3):
        d = np.array([np.cos(th[i]), np.sin(th[i]), 0.])
        rp = cfg.DRONES_INIT[i] + sp[i]*d*rt[i]; dp = rp + sp[i]*d*dl[i]; dp[2] -= 0.5*9.8*dl[i]**2
        rw = i+2
        ws.cell(rw,1,names[i]); ws.cell(rw,2,round(th[i],6)); ws.cell(rw,3,round(np.degrees(th[i]),4))
        ws.cell(rw,4,round(sp[i],2)); ws.cell(rw,5,round(rt[i],4)); ws.cell(rw,6,round(dl[i],4))
        ws.cell(rw,7,round(rp[0],2)); ws.cell(rw,8,round(rp[1],2)); ws.cell(rw,9,round(rp[2],2))
        ws.cell(rw,10,round(dp[0],2)); ws.cell(rw,11,round(dp[1],2)); ws.cell(rw,12,round(dp[2],2))
        if i==0: ws.cell(rw,13,round(f,4))
    wb.save("result2.xlsx")
    print("已保存 result2.xlsx")
    return x, f


def run_problem5():
    """问题5: 五机多弹 -> result3.xlsx"""
    print("\n" + "="*60)
    print("问题5: 5机协同多弹 (PSO分步)")
    print("="*60)

    order = cfg.INTERCEPT_ORDER
    n_d = 5; n_b = 3

    # Stage 1: per-drone optimization
    print("\n阶段1: 各无人机单独优化...")
    cfg_ranges = [
        {'th':(0.7,1.6), 'sp':(80,120), 'rt':(0,5), 'dl':(0,8)},
        {'th':(0.2,0.8), 'sp':(80,120), 'rt':(0,10), 'dl':(0,12)},
        {'th':(-0.8,-0.05), 'sp':(80,120), 'rt':(0,12), 'dl':(0,12)},
        {'th':(0,0.8), 'sp':(80,120), 'rt':(0,15), 'dl':(0,15)},
        {'th':(-0.5,0.1), 'sp':(80,120), 'rt':(0,15), 'dl':(0,15)},
    ]

    single = []
    for di in range(n_d):
        c = cfg_ranges[di]
        b = [c['th'], c['sp'], c['rt'], (1,5), (1,5), c['dl'], c['dl'], c['dl']]
        od = order[['FY1','FY2','FY3','FY4','FY5'][di]]

        def mk_obj(di, od):
            return lambda x: simulate_multi_bomb_single_drone(
                cfg.DRONES_INIT[di], x[0], x[1],
                np.array([x[2], x[2]+x[3], x[2]+x[3]+x[4]]),
                np.array([x[5], x[6], x[7]]),
                np.array(od), get_target_keypoints(360,5), cfg.DT, 25.0)

        pso = PSO(mk_obj(di, od), b, n_particles=150, max_iter=80, maximize=True, verbose=False)
        xo, fo = pso.optimize()
        single.append({'th':xo[0], 'sp':xo[1],
                       'rt':np.array([xo[2], xo[2]+xo[3], xo[2]+xo[3]+xo[4]]),
                       'dl':np.array([xo[5], xo[6], xo[7]]), 't':fo})
        print(f"  {['FY1','FY2','FY3','FY4','FY5'][di]}: {fo:.4f}s")

    # Stage 2: joint refinement
    print("\n阶段2: 联合微调...")
    bj = []
    dlt = 0.15
    for di in range(n_d):
        r = single[di]; c = cfg_ranges[di]
        bj.append((max(c['th'][0], r['th']-0.1), min(c['th'][1], r['th']+0.1)))
        bj.append((max(c['sp'][0], r['sp']-10), min(c['sp'][1], r['sp']+10)))
    for di in range(n_d):
        r = single[di]
        bj.append((max(0, r['rt'][0]-dlt), r['rt'][0]+dlt))
        bj.append((max(1, r['rt'][1]-r['rt'][0]-dlt), r['rt'][1]-r['rt'][0]+dlt))
        bj.append((max(1, r['rt'][2]-r['rt'][1]-dlt), r['rt'][2]-r['rt'][1]+dlt))
    for di in range(n_d):
        r = single[di]; c = cfg_ranges[di]
        for j in range(3):
            bj.append((max(c['dl'][0], r['dl'][j]-1), min(c['dl'][1], r['dl'][j]+1)))

    def obj_joint(x):
        idx = 0; dps = []
        for di in range(n_d):
            th=x[idx]; idx+=1; sp=x[idx]; idx+=1
            r1=x[idx]; idx+=1; i2=x[idx]; idx+=1; i3=x[idx]; idx+=1
            rt=np.array([r1, r1+i2, r1+i2+i3])
            dl=np.array([x[idx+j] for j in range(3)]); idx+=3
            od=order[['FY1','FY2','FY3','FY4','FY5'][di]]
            dps.append({'drone_init':cfg.DRONES_INIT[di],'theta':th,'speed':sp,
                        'release_times':rt,'detonation_delays':dl,'missile_indices':od})
        tt,_=simulate_multi_drone_multi_bomb(dps, cfg.DT, 35.0)
        return tt

    pso2 = PSO(obj_joint, bj, n_particles=300, max_iter=150, maximize=True, verbose=True)
    xj, fj = pso2.optimize()

    # Parse results
    idx = 0; final = []
    names=['FY1','FY2','FY3','FY4','FY5']
    print(f"\n联合优化总遮蔽时长: {fj:.4f} s")
    for di in range(n_d):
        th=xj[idx]; idx+=1; sp=xj[idx]; idx+=1
        r1=xj[idx]; idx+=1; i2=xj[idx]; idx+=1; i3=xj[idx]; idx+=1
        rt=np.array([r1,r1+i2,r1+i2+i3])
        dl=np.array([xj[idx+j] for j in range(3)]); idx+=3
        od=order[names[di]]
        d=np.array([np.cos(th), np.sin(th), 0.])
        print(f"\n  {names[di]}: θ={np.degrees(th):.1f}°, v={sp:.1f}m/s, 拦截:{[f'M{k+1}' for k in od]}")
        for j in range(3):
            dp = cfg.DRONES_INIT[di]+sp*d*rt[j]+sp*d*dl[j]; dp[2]-=0.5*9.8*dl[j]**2
            print(f"    弹{j+1}: 投放t={rt[j]:.4f}s, 延时={dl[j]:.4f}s, 起爆({dp[0]:.0f},{dp[1]:.0f},{dp[2]:.0f})")
        final.append({'th':th,'sp':sp,'rt':rt,'dl':dl,'od':od})

    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "问题5"
    h = ["无人机","航向角rad","航向角°","速度m/s","弹编号","目标导弹","投放时间s","延时s",
         "投放X","投放Y","投放Z","起爆X","起爆Y","起爆Z","总时长s"]
    for c,hv in enumerate(h,1): ws.cell(1,c,hv)
    rw=2
    for di in range(n_d):
        r=final[di]; d=np.array([np.cos(r['th']), np.sin(r['th']), 0.])
        for j in range(3):
            dp=cfg.DRONES_INIT[di]+r['sp']*d*r['rt'][j]+r['sp']*d*r['dl'][j]; dp[2]-=0.5*9.8*r['dl'][j]**2
            rp=cfg.DRONES_INIT[di]+r['sp']*d*r['rt'][j]
            ws.cell(rw,1,names[di]); ws.cell(rw,2,round(r['th'],6))
            ws.cell(rw,3,round(np.degrees(r['th']),4)); ws.cell(rw,4,round(r['sp'],2))
            ws.cell(rw,5,j+1); ws.cell(rw,6,f"M{r['od'][j]+1}")
            ws.cell(rw,7,round(r['rt'][j],4)); ws.cell(rw,8,round(r['dl'][j],4))
            ws.cell(rw,9,round(rp[0],2)); ws.cell(rw,10,round(rp[1],2)); ws.cell(rw,11,round(rp[2],2))
            ws.cell(rw,12,round(dp[0],2)); ws.cell(rw,13,round(dp[1],2)); ws.cell(rw,14,round(dp[2],2))
            if di==0 and j==0: ws.cell(rw,15,round(fj,4))
            rw+=1
    wb.save("result3.xlsx")
    print("已保存 result3.xlsx")
    return final, fj


if __name__ == "__main__":
    t0 = time.time()
    r1 = run_problem1()
    t1 = time.time()
    r2 = run_problem2()
    t2 = time.time()
    r3 = run_problem3()
    t3 = time.time()
    r4 = run_problem4()
    t4 = time.time()
    r5 = run_problem5()
    t5 = time.time()

    print("\n" + "="*70)
    print("                    最终结果汇总 (C题)")
    print("="*70)
    print(f"  问题1: {r1:.4f} s  [{t1-t0:.1f}s]")
    print(f"  问题2: {r2[1]:.4f} s  [{t2-t1:.1f}s]")
    print(f"  问题3: {r3[1]:.4f} s  [{t3-t2:.1f}s] -> result1.xlsx")
    print(f"  问题4: {r4[1]:.4f} s  [{t4-t3:.1f}s] -> result2.xlsx")
    print(f"  问题5: {r5[1]:.4f} s  [{t5-t4:.1f}s] -> result3.xlsx")
    print(f"  总耗时: {t5-t0:.1f}s ({(t5-t0)/60:.1f}min)")
    print("="*70)

    print("\n参考值 (2025 A题): 1.3915 / 4.5960 / 7.6500 / 11.7540 / 38.0600")
    print("C题参数: 下沉2.5m/s, 速度80-120m/s, P1投放1.2s/延时3.2s")
